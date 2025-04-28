# -*- coding: utf-8 -*-
"""
Braindog cuesheets - Extractor Mejorado para CUE Sheets desde Excel (.xlsx) o Markdown (.md).

Procesa archivos .xlsx o .md que contienen información de música (CUE sheets),
extrae los datos relevantes, genera reportes individuales por episodio en formato Markdown,
y un reporte global consolidado (en formato Markdown Y PDF) con estadísticas detalladas y gráficos.
Todos los tiempos se muestran en formato MM:SS.

Permite seleccionar archivos individuales o un directorio completo.
Permite especificar el directorio de salida para los reportes.
Permite especificar un nombre base para los reportes globales (--report-name).
Ofrece una interfaz gráfica opcional (--gui) para facilitar su uso.

Requiere: openpyxl, matplotlib, fpdf2, Pillow (para gráficos y logo en PDF), ttkbootstrap (opcional, para --gui)
Instalar con:
pip install openpyxl matplotlib fpdf2 Pillow
pip install ttkbootstrap  # Si quieres usar la interfaz gráfica

Historial de Cambios Recientes:
- v1.8.0 (Solicitud Usuario):
    - GUI: Duplicado el factor de escalado de la interfaz gráfica (tk scaling a 2.0) para aumentar tamaño de ventana y fuentes.
    - GUI: Título de la ventana cambiado a "Braindog cuesheets".
    - PDF: Aumentado el tamaño del logo en el pie de página (LOGO_WIDTH a 40).
    - PDF: Ajustada la paleta de colores (COLOR_PALETTE, CHART_COLORS) para coincidir con la del PDF de ejemplo proporcionado.
    - PDF: El fondo de los valores en la tabla de Resumen General ahora usa el color primario rosado, como en el PDF de ejemplo.
    - General: Actualizado nombre por defecto de reportes a "Braindog_Cuesheets_Report".
    - General: Actualizada descripción de la aplicación en CLI y comentarios.
    - Código completo proporcionado debido a cambios en múltiples áreas.
- v1.7.2 (Solicitud Usuario):
    - PDF: Reordenada sección "Análisis por Editora" para que aparezca después del "Resumen General".
    - PDF: Añadidos saltos de página automáticos antes de cada sección principal de análisis.
    - PDF Gráfico Circular (Publishers): Modificado para usar colormap ('tab20').
    - PDF Logo: Revisada la lógica.
- v1.7.1 (Solicitud Usuario):
    - PDF Página 1: Limitada a Título, Lista Episodios, Resumen General.
    - Resumen General PDF: Añadidas métricas específicas para RHAPSOLODY MUSIC LB.
    - Revisión de anchos de columna en tablas PDF.
- v1.7.0:
    - Alineación superior forzada en cabeceras de tablas PDF.
    - Lógica "Keep-Together" mejorada en PDF.
    - Añadido logo en pie de página PDF.
    - Rediseño visual de "Resumen General" en PDF (formato cajas).
- v1.6.0:
    - Paleta de colores personalizada para gráficos.
    - Opción --report-name / GUI.
    - Mejoras en la lógica "keep-together" PDF.
    - Revisión alineación encabezados tablas PDF.
- v1.5.1: Mejoras significativas en el diseño y formato del reporte PDF.
"""

import openpyxl
import re
import os
import argparse
from pathlib import Path
import sys
from collections import Counter, defaultdict # defaultdict puede ser útil
import datetime
from typing import List, Dict, Any, Tuple, Optional
import math
import shutil
import traceback # Para imprimir errores detallados

# --- Importaciones para Gráficos ---
MATPLOTLIB_AVAILABLE = False
try:
    import matplotlib
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
    print("INFO: Matplotlib encontrado. Se generarán gráficos.")
except ImportError:
    print("ADVERTENCIA: matplotlib no está instalado. No se generarán gráficos.", file=sys.stderr)
    print("Instala con: pip install matplotlib", file=sys.stderr)

# --- Importaciones para PDF ---
FPDF2_AVAILABLE = False
try:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos # Necesario para reemplazar ln=
    FPDF2_AVAILABLE = True
    print("INFO: FPDF2 encontrado. Se generará reporte global en PDF.")
except ImportError:
    print("ADVERTENCIA: fpdf2 no está instalado. No se generará el reporte global en PDF.", file=sys.stderr)
    print("Instala con: pip install fpdf2", file=sys.stderr)

# --- Importaciones para Imágenes/Logo en PDF ---
PIL_AVAILABLE = False
try:
    from PIL import Image
    PIL_AVAILABLE = True
    print("INFO: Pillow (PIL) encontrado. Se podrán insertar gráficos y logo en PDF.")
except ImportError:
     print("ADVERTENCIA: Pillow (PIL) no está instalado. No se podrán insertar gráficos ni logo en los reportes PDF.", file=sys.stderr)
     print("Instala con: pip install Pillow", file=sys.stderr)


# --- Importaciones para GUI (si se usa --gui) ---
GUI_ENABLED = False
try:
    import ttkbootstrap as ttkb
    from tkinter import filedialog, messagebox
    from ttkbootstrap.scrolled import ScrolledText
    GUI_ENABLED = True
except ImportError:
    if '--gui' in sys.argv: # Solo mostrar advertencia si se intenta usar la GUI
        print("Advertencia: ttkbootstrap no está instalado. La interfaz gráfica (--gui) no estará disponible.", file=sys.stderr)
        print("Instala con: pip install ttkbootstrap", file=sys.stderr)


# =========================
# Configuración Centralizada
# =========================
COL_TITULO: int = 4
COL_TIEMPO: int = 7
COL_COMPOSITOR: int = 9
COL_PUBLISHER: int = 14
ROW_START: int = 17 # Fila donde empiezan los datos en Excel (1-based)
REGEX_EPISODIO: str = r'(?:EP|CAP|Episodio)\s*(\d+)' # Regex para encontrar el número de episodio
DEFAULT_EPISODIO: str = "000" # Episodio por defecto si no se encuentra
REGEX_MARKDOWN_ROW: str = r'^\s*\|\s*(\d+)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|' # Regex para filas de tabla MD
# <<< CAMBIO v1.8.0: Nombre por defecto actualizado >>>
DEFAULT_REPORT_NAME: str = "Braindog_Cuesheets_Report" # Nombre base por defecto para reportes globales
REPORTE_GLOBAL_MD_FILENAME_FORMAT: str = "{report_name}_Global.md"
REPORTE_GLOBAL_PDF_FILENAME_FORMAT: str = "{report_name}_Global.pdf"
REPORTE_EPISODIO_FILENAME_FORMAT: str = "Reporte_Musical_Episodio_{episodio}.md"
MARKDOWN_EPISODIO_FILENAME_FORMAT: str = "Episodio_{episodio}.md" # Para tabla simple si entrada es Excel
CHARTS_SUBDIR: str = "charts" # Subdirectorio para guardar gráficos
LOGO_FILENAME: str = "logo.png" # Nombre del archivo del logo
TOP_N_PISTAS_GLOBAL: int = 20
TOP_N_COMPOSITOR_GLOBAL: int = 15
TOP_N_PUBLISHER_GLOBAL: int = 15
TOP_N_PISTAS_DETALLE_GLOBAL: int = 10
TOP_N_PISTAS_EPISODIO: int = 10
TOP_N_COMPOSITOR_EPISODIO: int = 10
TOP_N_PUBLISHER_EPISODIO: int = 10
PIE_CHART_OTHERS_THRESHOLD: float = 3.0 # Porcentaje para agrupar en 'Otros' en pie chart
BAR_CHART_TOP_N_COMPOSERS: int = 15 # Top N para gráfico de barras de compositores
BAR_CHART_TOP_N_TRACKS_TIME: int = 15 # Top N para gráfico de barras de pistas por tiempo
PUBLISHER_RHAPSODY = "RHAPSOLODY MUSIC LB" # Constante para el nombre de Rhapsody

# --- Paleta de Colores Actualizada (v1.8.0) ---
# Basada en el PDF de ejemplo
COLOR_PALETTE = {
    'primary_pink': '#FF5A78',    # Rosado principal (Gráficos, Fondo Valor Resumen)
    'medium_purple': '#8A4D76',   # Púrpura medio (Títulos H2, Cabeceras Tabla)
    # 'dark_purple': '#4B003A',     # Púrpura oscuro (No claramente usado en PDF para texto, mantener para posible uso)
    'lighter_pink': '#FFAAB9',    # Rosado claro (Alternativa o para otros elementos)
    # 'contrast_accent': '#FFD166', # Amarillo/Dorado (No visto en PDF, mantener para posible uso)
    'neutral_gray': '#B0B0B0',    # Gris neutro (Bordes, Ejes Gráfico, Texto Secundario)
    'light_gray_bg': '#F5F5F5',   # Gris muy claro (Fondo Métrica Resumen, Zebra Tablas)
    'text_on_dark': '#FFFFFF',    # Blanco (Texto sobre fondos oscuros/púrpura)
    'text_on_light': '#333333',   # Gris oscuro/Negro (Texto principal, Títulos H1)
}

# Asignaciones específicas para gráficos (Actualizado v1.8.0)
CHART_COLORS = {
    # Colores para Pie Chart (intentará usar colormap si necesita más)
    'pie': [COLOR_PALETTE['primary_pink'], COLOR_PALETTE['medium_purple'], '#EAB3E1', COLOR_PALETTE['neutral_gray'], '#FFC7B3'], # Añadidos tonos intermedios
    # Colores para Barras (principalmente el rosado del PDF)
    'bar_composers': COLOR_PALETTE['primary_pink'],
    'bar_tracks': COLOR_PALETTE['medium_purple'], # Usar púrpura para diferenciar
    'bar_episodes_minutes': COLOR_PALETTE['primary_pink'], # Rosado para minutos
    'bar_episodes_tracks': COLOR_PALETTE['medium_purple'], # Púrpura para pistas únicas
    # Colores de Texto y Ejes
    'axis_labels': COLOR_PALETTE['text_on_light'],
    'titles': COLOR_PALETTE['text_on_light'],
    'grid': COLOR_PALETTE['neutral_gray']
}


# =========================
# Funciones Auxiliares
# =========================

def formatear_tiempo(segundos_totales: int) -> str:
    """
    Convierte segundos a formato MM:SS.
    Si los segundos son None, devuelve "00:00". Maneja negativos tratándolos como 0.
    Los minutos pueden ser > 59 si el total excede 1 hora.
    """
    if segundos_totales is None or segundos_totales < 0:
        segundos_totales = 0

    minutos_totales = segundos_totales // 60
    segundos_restantes = segundos_totales % 60

    # Formato siempre MM:SS, asegurando dos dígitos para ambos
    return f"{minutos_totales:02d}:{segundos_restantes:02d}"

def time_formatter(x: float, pos: Optional[int] = None) -> str:
    """ Formateador para ejes Matplotlib usando MM:SS. """
    return formatear_tiempo(int(x))

def parsear_y_formatear_tiempo(valor_celda_tiempo: Any) -> Tuple[str, int]:
    """
    Parsea varios formatos de tiempo y devuelve (MM:SS str, segundos int).
    El formato de salida de la cadena siempre será MM:SS.
    Redondea hacia arriba si hay fracciones de segundo.
    """
    if valor_celda_tiempo is None:
        return "00:00", 0 # Devuelve formato MM:SS por defecto

    tiempo_str = str(valor_celda_tiempo).strip()
    segundos_totales = 0
    fraccion_presente = False

    try:
        # 1. Formato HH:MM:SS;ff (Excel con frames/ff) o HH:MM:SS.ff
        match_ff = re.search(r'(\d{1,2}):(\d{1,2}):(\d{1,2})[;.:](\d+)', tiempo_str)
        if match_ff:
            h, m, s, f = map(int, match_ff.groups())
            segundos_totales = h * 3600 + m * 60 + s
            fraccion_presente = f > 0 # Cualquier frame/fracción > 0 cuenta
        else:
            # 2. Formato HH:MM:SS o HH:MM:SS.sss (con milisegundos o decimal)
            match_hms_decimal = re.search(r'(\d{1,2}):(\d{1,2}):(\d+(\.\d+)?)', tiempo_str) # Segundos pueden tener decimal
            if match_hms_decimal:
                h = int(match_hms_decimal.group(1))
                m = int(match_hms_decimal.group(2))
                s_float = float(match_hms_decimal.group(3))
                s_int = int(s_float)
                segundos_totales = h * 3600 + m * 60 + s_int
                fraccion_presente = s_float > s_int # Si la parte flotante es mayor que el entero
            else:
                # 3. Formato MM:SS o MM:SS.sss
                match_ms_decimal = re.search(r'(\d{1,3}):(\d{1,2}(\.\d+)?)', tiempo_str) # Minutos > 59 permitidos, segundos con decimal
                if match_ms_decimal:
                    m = int(match_ms_decimal.group(1))
                    s_float = float(match_ms_decimal.group(2))
                    s_int = int(s_float)
                    segundos_totales = m * 60 + s_int
                    fraccion_presente = s_float > s_int
                else:
                     # 4. Formato numérico de Excel (fracción de día)
                     if isinstance(valor_celda_tiempo, (int, float)):
                         # Evitar conversión si el número es muy grande (podría ser segundos ya)
                         if valor_celda_tiempo > 5: # Heurística: >5 probablemente no es fracción de día
                             segundos_totales = int(valor_celda_tiempo)
                             fraccion_presente = (valor_celda_tiempo - segundos_totales) > 1e-9
                         else:
                             segundos_float = valor_celda_tiempo * 86400 # 24 * 60 * 60
                             segundos_totales = int(segundos_float)
                             fraccion_presente = (segundos_float - segundos_totales) > 1e-9
                     # 5. Objeto datetime.time de Python
                     elif isinstance(valor_celda_tiempo, datetime.time):
                         segundos_totales = valor_celda_tiempo.hour * 3600 + valor_celda_tiempo.minute * 60 + valor_celda_tiempo.second
                         fraccion_presente = valor_celda_tiempo.microsecond > 0
                     # 6. Entero interpretado como segundos
                     elif isinstance(valor_celda_tiempo, int):
                          segundos_totales = valor_celda_tiempo
                          fraccion_presente = False # Asumimos que no hay fracción si es entero
                     else:
                        # Si no coincide con ningún formato conocido
                        raise ValueError(f"Formato de tiempo no reconocido: {tiempo_str}")

    except Exception as e:
        print(f"Advertencia: No se pudo parsear tiempo '{tiempo_str}'. Usando 0s. Error: {e}", file=sys.stderr)
        return "00:00", 0 # Devuelve formato MM:SS

    # Redondear hacia arriba si hay fracción de segundo (y no es negativo)
    if fraccion_presente and segundos_totales >= 0:
        segundos_totales += 1

    # Asegurarse de que el resultado final no sea negativo
    if segundos_totales < 0:
        segundos_totales = 0

    # Llama a la función formatear_tiempo actualizada que devuelve MM:SS
    return formatear_tiempo(segundos_totales), segundos_totales


def limpiar_participante(valor_celda: Any) -> str:
    """ Limpia nombres de compositores/publishers. """
    if not valor_celda: return "N/A"
    if not isinstance(valor_celda, str): valor_celda = str(valor_celda)
    # Elimina texto entre paréntesis (ej. P.R.O.)
    texto_limpio = re.sub(r'\([^)]*\)', '', valor_celda).strip()
    # Divide por '/' y limpia espacios de cada parte
    participantes = [p.strip() for p in texto_limpio.split('/') if p.strip()]
    # Une con ' / ' si hay más de uno, o devuelve el único, o N/A si queda vacío
    return ' / '.join(participantes) if participantes else "N/A"

# =========================
# Clase Auxiliar para PDF (Actualizada v1.8.0 con colores y tamaño logo)
# =========================
if FPDF2_AVAILABLE:
    class PDFReport(FPDF):
        # --- CONSTANTES DE ESTILO PDF (Actualizado v1.8.0) ---
        COLOR_PRIMARY_TEXT = tuple(int(COLOR_PALETTE['text_on_light'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        COLOR_SECONDARY_TEXT = tuple(int(COLOR_PALETTE['neutral_gray'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        COLOR_LINK = (0, 0, 255) # Azul estándar para links
        COLOR_TABLE_HEADER_BG = tuple(int(COLOR_PALETTE['medium_purple'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Púrpura medio
        COLOR_TABLE_HEADER_TEXT = tuple(int(COLOR_PALETTE['text_on_dark'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Blanco
        COLOR_TABLE_BORDER = tuple(int(COLOR_PALETTE['neutral_gray'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Gris neutro
        COLOR_ZEBRA_STRIPE = tuple(int(COLOR_PALETTE['light_gray_bg'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Gris claro
        # Colores para Resumen General estilo cajas (Actualizado v1.8.0)
        COLOR_RESUMEN_BG_METRIC = tuple(int(COLOR_PALETTE['light_gray_bg'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Fondo métrica: Gris claro
        COLOR_RESUMEN_BG_VALUE = tuple(int(COLOR_PALETTE['primary_pink'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) # Fondo valor: Rosado primario

        FONT_SIZE_H1 = 18
        FONT_SIZE_H2 = 14
        FONT_SIZE_H3 = 12
        FONT_SIZE_BODY = 10
        FONT_SIZE_TABLE_HEADER = 9
        FONT_SIZE_TABLE_CELL = 8
        FONT_SIZE_FOOTER = 8
        FONT_SIZE_RESUMEN = 9

        LINE_HEIGHT_MULTIPLIER = 1.35
        SPACING_BEFORE_TITLE = 8
        SPACING_AFTER_TITLE = 4
        SPACING_AFTER_TABLE = 6
        SPACING_AFTER_CHART = 6
        SPACING_AFTER_TEXT = 3
        SPACING_AFTER_RESUMEN_ROW = 0.5

        ESTIMATED_TITLE_H3_HEIGHT = 8
        ESTIMATED_CHART_HEIGHT = 70
        ESTIMATED_TABLE_HEADER_HEIGHT = 8
        ESTIMATED_TABLE_ROW_HEIGHT = 6
        KEEP_TOGETHER_MIN_ROWS = 2

        # <<< CAMBIO v1.8.0: Tamaño del logo aumentado >>>
        LOGO_WIDTH = 40 # Ancho del logo en mm (era 25)

        def __init__(self, report_name: str = DEFAULT_REPORT_NAME, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.report_name = report_name # Guardar nombre del reporte
            self.alias_nb_pages()
            self.chart_paths: Dict[str, str] = {}
            self.logo_path: Optional[str] = None
            # Intentar encontrar el logo en el directorio del script
            try:
                script_dir = Path(__file__).parent.resolve()
                potential_logo_path = script_dir / LOGO_FILENAME
                if potential_logo_path.exists():
                    self.logo_path = str(potential_logo_path)
                    print(f"INFO: Logo encontrado en: {self.logo_path}")
                else:
                    print(f"ADVERTENCIA: No se encontró el archivo de logo '{LOGO_FILENAME}' en {script_dir}. El logo no se incluirá en el PDF.", file=sys.stderr)
            except Exception as e:
                 print(f"ADVERTENCIA: Error al intentar localizar el logo: {e}. El logo no se incluirá.", file=sys.stderr)


            self.WIDTH_A4 = 210
            self.HEIGHT_A4 = 297
            self.MARGIN_LEFT = 15
            self.MARGIN_RIGHT = 15
            self.MARGIN_TOP = 15
            self.MARGIN_BOTTOM = 20 # Margen inferior suficiente para logo más grande
            self.set_margins(self.MARGIN_LEFT, self.MARGIN_TOP, self.MARGIN_RIGHT)
            self.set_auto_page_break(auto=True, margin=self.MARGIN_BOTTOM)

            # Carga de fuentes
            try:
                font_dir = Path(__file__).parent.resolve()
                try:
                    if "dejavu" not in self.fonts:
                         self.add_font('DejaVu', '', os.path.join(font_dir, 'DejaVuSans.ttf'))
                         self.add_font('DejaVu', 'B', os.path.join(font_dir, 'DejaVuSans-Bold.ttf'))
                    self.DEFAULT_FONT = 'DejaVu'
                    print("INFO: Usando fuente DejaVu para PDF (soporta UTF-8).")
                except RuntimeError:
                     raise FileNotFoundError("Las fuentes DejaVu (DejaVuSans.ttf, DejaVuSans-Bold.ttf) no se encontraron o no se pudieron cargar desde el directorio del script.")
            except Exception as e:
                self.DEFAULT_FONT = 'Helvetica'
                print(f"ADVERTENCIA: Fuente DejaVu no encontrada o error al cargar ({e}). Usando '{self.DEFAULT_FONT}'.", file=sys.stderr)
                print("-> Para mejor soporte de caracteres, descarga DejaVuSans.ttf y DejaVuSans-Bold.ttf y colócalas junto al script.", file=sys.stderr)

            self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_BODY)
            self.set_text_color(*self.COLOR_PRIMARY_TEXT)

        def header(self):
            pass # Sin cabecera por defecto

        def footer(self):
            # Dibuja número de página centrado
            y_pos_footer_text = -self.MARGIN_BOTTOM + 5
            self.set_y(y_pos_footer_text)
            self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_FOOTER)
            self.set_text_color(*self.COLOR_SECONDARY_TEXT)
            self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', border=0, align='C')
            self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Restaurar color

            # Añadir Logo (más grande - v1.8.0)
            if self.logo_path and PIL_AVAILABLE:
                try:
                    with Image.open(self.logo_path) as img:
                        dpi_x, dpi_y = img.info.get('dpi', (96, 96))
                        dpi = max(dpi_x, dpi_y, 96)
                        img_w_px, img_h_px = img.size
                    if img_w_px <= 0 or img_h_px <= 0: raise ValueError("Dimensiones de imagen inválidas.")

                    px_to_mm = 25.4 / dpi
                    logo_original_w_mm = img_w_px * px_to_mm
                    logo_original_h_mm = img_h_px * px_to_mm

                    # <<< CAMBIO v1.8.0: Usa self.LOGO_WIDTH que ahora es 40 >>>
                    logo_w = self.LOGO_WIDTH
                    aspect_ratio = logo_original_h_mm / logo_original_w_mm if logo_original_w_mm > 0 else 1
                    logo_h = logo_w * aspect_ratio

                    logo_x = self.WIDTH_A4 - self.MARGIN_RIGHT - logo_w
                    # Ajustar Y para que quepa el logo más grande
                    logo_y = self.HEIGHT_A4 - self.MARGIN_BOTTOM - logo_h + 2 # Posicionar desde abajo, un poco más arriba
                    if logo_y < self.t_margin: # Seguridad, si el logo es gigantesco
                        logo_y = self.t_margin

                    current_x = self.get_x()
                    current_y = self.get_y()
                    self.image(self.logo_path, x=logo_x, y=logo_y, w=logo_w, h=logo_h)
                    self.set_xy(current_x, current_y)
                except FileNotFoundError: pass # Advertencia ya dada en init
                except Exception as e: print(f"Error al insertar logo en PDF: {e}", file=sys.stderr)

        def chapter_title(self, title: str, level: int = 1):
            # Ajusta espacio antes del título
            self.ln(self.SPACING_BEFORE_TITLE if self.get_y() > self.t_margin + 10 else 2)

            if level == 1: # Título Principal (Reporte Global)
                self.set_font(self.DEFAULT_FONT, 'B', self.FONT_SIZE_H1)
                # <<< CAMBIO v1.8.0: Usar color primario de texto (oscuro) como en PDF ejemplo >>>
                self.set_text_color(*self.COLOR_PRIMARY_TEXT)
                self.multi_cell(0, 10, title, border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                self.ln(self.SPACING_AFTER_TITLE * 1.5)
            elif level == 2: # Títulos de Sección Principal
                self.set_font(self.DEFAULT_FONT, 'B', self.FONT_SIZE_H2)
                # <<< CAMBIO v1.8.0: Usar púrpura medio como en PDF ejemplo >>>
                self.set_text_color(*tuple(int(COLOR_PALETTE['medium_purple'].lstrip('#')[i:i+2], 16) for i in (0, 2, 4)))
                self.multi_cell(0, 9, title, border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                self.ln(self.SPACING_AFTER_TITLE)
            else: # H3 (Subtítulos para Tablas/Gráficos)
                self.set_font(self.DEFAULT_FONT, 'B', self.FONT_SIZE_H3)
                self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Texto normal oscuro
                self.multi_cell(0, 8, title, border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
                self.ln(self.SPACING_AFTER_TITLE * 0.5)

            self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Asegurar color primario después

        def body_text(self, text: str):
            self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_BODY)
            self.multi_cell(0, self.font_size * self.LINE_HEIGHT_MULTIPLIER, text, border=0, align='L', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.ln(self.SPACING_AFTER_TEXT)

        def bullet_item(self, text: str):
             self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_BODY)
             try: bullet = chr(149) if self.DEFAULT_FONT == 'DejaVu' else '-'
             except: bullet = '*'
             self.cell(5) # Sangría
             self.multi_cell(0, self.font_size * self.LINE_HEIGHT_MULTIPLIER, f"{bullet} {text}", border=0, align='L', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
             self.ln(1)

        def _get_estimated_element_height(self, element_type: str, content: Any = None) -> float:
            """ Estima la altura de diferentes elementos para keep-together. """
            if element_type == 'title_h3':
                return self.FONT_SIZE_H3 / 2.5 + self.SPACING_AFTER_TITLE * 0.5
            elif element_type == 'table_header':
                 return self.FONT_SIZE_TABLE_HEADER / 2.5 + 2
            elif element_type == 'table_row':
                return self.FONT_SIZE_TABLE_CELL / 2.5 + 2
            elif element_type == 'chart':
                return self.ESTIMATED_CHART_HEIGHT + self.SPACING_AFTER_CHART
            elif element_type == 'resumen_row':
                 return self.FONT_SIZE_RESUMEN / 2.5 + self.SPACING_AFTER_RESUMEN_ROW
            return 5.0

        def _check_keep_together(self, required_height: float, reason: str = "element"):
            """ Comprueba si la altura requerida cabe, si no, salta de página. """
            available_space = self.page_break_trigger - self.get_y()
            if required_height > available_space:
                self.add_page()
                return True
            return False

        # MÉTODO MEJORADO: Permite añadir filas de datos extra y usa colores actualizados (v1.8.0)
        def add_resumen_general(self, data: List[List[str]], title: Optional[str] = None, extra_data: Optional[List[List[str]]] = None):
            """
            Añade la sección Resumen General con estilo de cajas (gris claro + rosado primario).
            Permite añadir un bloque 'extra_data' debajo del principal.
            """
            if title:
                 self.chapter_title(title, level=2) # Usa H2 (púrpura)

            all_data = data + (extra_data if extra_data else [])

            estimated_h2_height = self.FONT_SIZE_H2 / 2.5 + self.SPACING_AFTER_TITLE
            required_height = estimated_h2_height + sum(self._get_estimated_element_height('resumen_row') for _ in all_data)
            self._check_keep_together(required_height, "Resumen General completo")

            self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_RESUMEN)
            line_height = self.font_size * self.LINE_HEIGHT_MULTIPLIER * 1.2
            available_width = self.w - self.l_margin - self.r_margin
            metric_width = available_width * 0.65
            value_width = available_width * 0.35

            for i, (metric, value) in enumerate(all_data):
                 start_y = self.get_y()
                 # --- Dibujar celda Métrica (izquierda) ---
                 self.set_x(self.l_margin)
                 self.set_fill_color(*self.COLOR_RESUMEN_BG_METRIC) # Gris claro
                 self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Texto oscuro
                 # Negrita solo si NO es un compositor de Rhapsody
                 if not metric.startswith("  "):
                     self.set_font(self.DEFAULT_FONT, 'B', self.FONT_SIZE_RESUMEN)
                 else:
                     self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_RESUMEN)
                 self.multi_cell(metric_width, line_height, metric, border=1, align='L', fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP)

                 # --- Dibujar celda Valor (derecha) ---
                 self.set_xy(self.l_margin + metric_width, start_y)
                 # <<< CAMBIO v1.8.0: Fondo valor es rosado primario >>>
                 self.set_fill_color(*self.COLOR_RESUMEN_BG_VALUE)
                 self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Texto oscuro
                 self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_RESUMEN)
                 self.multi_cell(value_width, line_height, value, border=1, align='R', fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

                 if i < len(all_data) - 1:
                     self.ln(self.SPACING_AFTER_RESUMEN_ROW)

            self.ln(self.SPACING_AFTER_TABLE)

        def add_table(self, headers: List[str], data: List[List[str]], col_widths: Optional[List[float]] = None, title: Optional[str] = None):
            """ Añade una tabla con estilo (cabecera púrpura/blanco, zebra gris claro), keep-together. """
            required_height = 0
            if title: required_height += self._get_estimated_element_height('title_h3')
            required_height += self._get_estimated_element_height('table_header')
            num_rows_to_keep = min(len(data) if data else 0, self.KEEP_TOGETHER_MIN_ROWS)
            required_height += self._get_estimated_element_height('table_row') * num_rows_to_keep

            self._check_keep_together(required_height, f"table '{title or 'untitled'}'")

            if title: self.chapter_title(title, level=3) # H3 normal

            effective_width = self.w - self.l_margin - self.r_margin
            num_cols = len(headers)
            if col_widths is None:
                col_widths = [effective_width / num_cols] * num_cols
            elif abs(sum(col_widths) - effective_width) > 1:
                current_sum = sum(col_widths)
                scale_factor = effective_width / current_sum if current_sum > 0 else 0
                col_widths = [w * scale_factor for w in col_widths]

            base_header_line_height = self.FONT_SIZE_TABLE_HEADER / 2.5

            def draw_header():
                self.set_font(self.DEFAULT_FONT, 'B', self.FONT_SIZE_TABLE_HEADER)
                # <<< CAMBIO v1.8.0: Usa colores de tabla definidos (púrpura/blanco) >>>
                self.set_fill_color(*self.COLOR_TABLE_HEADER_BG)
                self.set_text_color(*self.COLOR_TABLE_HEADER_TEXT)
                self.set_draw_color(*self.COLOR_TABLE_BORDER) # Borde gris
                self.set_line_width(0.3)
                start_y_header = self.get_y()
                current_x_header = self.get_x()

                max_header_lines = 1
                for i, header in enumerate(headers):
                    if col_widths[i] <=0: continue
                    lines_needed = max(
                        math.ceil(self.get_string_width(line) / col_widths[i]) for line in header.split('\n')
                    ) if self.get_string_width(header) > 0 else 1
                    lines_needed = max(lines_needed, header.count('\n') + 1)
                    max_header_lines = max(max_header_lines, lines_needed)
                header_row_height = max_header_lines * base_header_line_height * self.LINE_HEIGHT_MULTIPLIER

                for i, header in enumerate(headers):
                    self.set_xy(current_x_header + sum(col_widths[:i]), start_y_header)
                    self.multi_cell(col_widths[i], base_header_line_height * self.LINE_HEIGHT_MULTIPLIER, header, border=1, align='C', fill=True, new_x=XPos.RIGHT, new_y=YPos.TOP, max_line_height=base_header_line_height * self.LINE_HEIGHT_MULTIPLIER)

                self.set_y(start_y_header + header_row_height)
                return header_row_height

            header_actual_height = draw_header()

            # Filas de Datos
            self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_TABLE_CELL)
            self.set_text_color(*self.COLOR_PRIMARY_TEXT) # Texto oscuro
            self.set_draw_color(*self.COLOR_TABLE_BORDER) # Borde gris
            self.set_line_width(0.2)
            base_cell_line_height = self.FONT_SIZE_TABLE_CELL / 2.5
            fill = False

            for row_idx, row in enumerate(data):
                max_lines_in_row = 1
                for i, cell_text in enumerate(row):
                    if col_widths[i] <=0: continue
                    cell_content = str(cell_text) if cell_text is not None else ""
                    lines_needed = max(
                         math.ceil(self.get_string_width(line) / col_widths[i]) for line in cell_content.split('\n')
                    ) if self.get_string_width(cell_content) > 0 else 1
                    lines_needed = max(lines_needed, cell_content.count('\n') + 1)
                    max_lines_in_row = max(max_lines_in_row, lines_needed)
                row_height_needed = max_lines_in_row * base_cell_line_height * self.LINE_HEIGHT_MULTIPLIER

                if self.get_y() + row_height_needed > self.page_break_trigger:
                     self.add_page()
                     header_actual_height = draw_header() # Redibujar cabecera
                     self.set_font(self.DEFAULT_FONT, '', self.FONT_SIZE_TABLE_CELL)
                     self.set_text_color(*self.COLOR_PRIMARY_TEXT)
                     self.set_draw_color(*self.COLOR_TABLE_BORDER)
                     self.set_line_width(0.2)
                     fill = False # Reset zebra

                fill = row_idx % 2 == 1
                if fill: self.set_fill_color(*self.COLOR_ZEBRA_STRIPE) # Zebra gris claro

                start_y_row = self.get_y()
                current_x_row = self.get_x()

                for i, cell_text in enumerate(row):
                    self.set_xy(current_x_row + sum(col_widths[:i]), start_y_row)
                    cell_content = str(cell_text) if cell_text is not None else ""
                    align = 'L'
                    if isinstance(cell_text, (int, float)) or \
                       re.fullmatch(r'-?[\d,.]+%?', cell_content) or \
                       re.fullmatch(r'\d+:\d{2}', cell_content) or \
                       (cell_content.isdigit() and not re.fullmatch(r'\d{1,3}', cell_content)):
                         align = 'R'
                    self.multi_cell(col_widths[i], base_cell_line_height * self.LINE_HEIGHT_MULTIPLIER, cell_content, border=1, align=align, fill=fill, new_x=XPos.RIGHT, new_y=YPos.TOP, max_line_height=base_cell_line_height * self.LINE_HEIGHT_MULTIPLIER)

                self.set_y(start_y_row + row_height_needed)

            self.ln(self.SPACING_AFTER_TABLE)


        def add_chart(self, chart_id: str, title: str):
            """ Inserta un gráfico con título H3 y keep-together. """
            required_height = self._get_estimated_element_height('title_h3') + self._get_estimated_element_height('chart')
            page_jumped = self._check_keep_together(required_height, f"chart '{title}'")

            self.chapter_title(title, level=3) # Título H3 normal

            chart_path_str = self.chart_paths.get(chart_id)
            if not chart_path_str:
                message = f"*Nota: No se pudo generar/encontrar gráfico '{chart_id}'." if MATPLOTLIB_AVAILABLE else f"*Nota: Gráficos omitidos (matplotlib no disponible).* "
                self.body_text(message)
                self.ln(self.SPACING_AFTER_CHART)
                return

            chart_path = Path(chart_path_str)

            if chart_path.exists() and PIL_AVAILABLE:
                try:
                    with Image.open(chart_path) as img:
                        dpi_x, dpi_y = img.info.get('dpi', (96, 96))
                        dpi = max(dpi_x, dpi_y, 96)
                        img_original_w_px, img_original_h_px = img.size
                    if img_original_w_px <= 0 or img_original_h_px <= 0: raise ValueError("Invalid image dimensions.")

                    px_to_mm = 25.4 / dpi
                    img_original_w_mm = img_original_w_px * px_to_mm
                    img_original_h_mm = img_original_h_px * px_to_mm

                    available_width = self.w - self.l_margin - self.r_margin
                    img_w = min(available_width * 0.95, img_original_w_mm)
                    aspect_ratio = img_original_h_mm / img_original_w_mm if img_original_w_mm > 0 else 1
                    img_h = img_w * aspect_ratio

                    space_needed_now = img_h + self.SPACING_AFTER_CHART
                    if self.get_y() + space_needed_now > self.page_break_trigger:
                         self.add_page()
                         self.chapter_title(title, level=3) # Redibujar título

                    x_pos = self.l_margin + (available_width - img_w) / 2
                    y_pos = self.get_y()
                    self.image(chart_path, x=x_pos, y=y_pos, w=img_w, h=img_h)
                    self.set_y(y_pos + img_h + self.SPACING_AFTER_CHART)

                except Exception as e:
                    print(f"Error al insertar imagen PDF '{chart_path}': {e}", file=sys.stderr)
                    self.body_text(f"*Error al procesar/insertar gráfico '{chart_path.name}'.*")
                    self.ln(self.SPACING_AFTER_CHART)
            else:
                 message = ""
                 if not chart_path.exists(): message = f"*Nota: No se encontró archivo gráfico '{chart_id}'.*"
                 elif not PIL_AVAILABLE: message = f"*Nota: No se insertó gráfico '{chart_path.name}' (Pillow no instalado).* "
                 self.body_text(message)
                 self.ln(self.SPACING_AFTER_CHART)


# =========================
# Funciones de Procesamiento y Estadísticas
# (Sin cambios significativos aquí)
# =========================
def calcular_estadisticas(datos_tabla: List[Dict[str, Any]]) -> Dict[str, Any]:
    """ Calcula estadísticas detalladas a partir de una lista de datos de pistas. """
    stats = {
        'total_pistas': len(datos_tabla),
        'duracion_total_segundos': 0,
        'compositores': Counter(),        # Contador de ocurrencias por compositor
        'compositores_tiempo': Counter(), # Contador de segundos totales por compositor
        'publishers': Counter(),          # Contador de ocurrencias por publisher
        'publishers_tiempo': Counter(),   # Contador de segundos totales por publisher
        'pistas_por_duracion': Counter(), # Contador de pistas por rangos de duración
        'titulos': Counter(),             # Contador de ocurrencias por título
        'episodios': set(),               # Conjunto de episodios únicos
        'pistas_consolidadas': [],        # Lista de pistas únicas (Título+Comp) agregadas, ordenada por tiempo
        'pistas_repetidas_detalle': [],   # Lista detallada de pistas únicas, ordenada por ocurrencias
        'unique_tracks_count': 0          # Cuenta total de pistas únicas (Título+Compositor)
    }
    pistas_agrupadas_temp = {} # Clave: "Título|Compositor"

    for pista in datos_tabla:
        stats['episodios'].add(pista.get('episode', DEFAULT_EPISODIO))
        titulo_limpio = pista.get('title', 'N/A').strip()
        stats['titulos'][titulo_limpio] += 1
        duracion_segundos = pista.get('duration_seconds', 0)
        stats['duracion_total_segundos'] += duracion_segundos

        compositor_str = pista.get('composer', 'N/A')
        if compositor_str != 'N/A':
            for comp in compositor_str.split(' / '):
                comp_limpio = comp.strip()
                if comp_limpio:
                    stats['compositores'][comp_limpio] += 1
                    stats['compositores_tiempo'][comp_limpio] += duracion_segundos

        publisher_str = pista.get('publisher', 'N/A')
        if publisher_str != 'N/A':
            for pub in publisher_str.split(' / '):
                pub_limpio = pub.strip()
                if pub_limpio:
                    stats['publishers'][pub_limpio] += 1
                    stats['publishers_tiempo'][pub_limpio] += duracion_segundos

        if duracion_segundos > 0:
            rango_idx = (duracion_segundos - 1) // 30
            rango_inicio_seg = rango_idx * 30 + 1
            rango_fin_seg = (rango_idx + 1) * 30
            rango_str = f"{formatear_tiempo(rango_inicio_seg)}-{formatear_tiempo(rango_fin_seg)}"
            stats['pistas_por_duracion'][rango_str] += 1

        clave_pista = f"{titulo_limpio}|{compositor_str}"
        episodio_actual = pista.get('episode', DEFAULT_EPISODIO)
        if clave_pista in pistas_agrupadas_temp:
            pistas_agrupadas_temp[clave_pista]['duration_seconds'] += duracion_segundos
            pistas_agrupadas_temp[clave_pista]['ocurrencias'] += 1
            pistas_agrupadas_temp[clave_pista]['episodios'].add(episodio_actual)
        else:
            pistas_agrupadas_temp[clave_pista] = {
                'title': titulo_limpio,
                'composer': compositor_str,
                'publisher': pista.get('publisher', 'N/A'),
                'duration_seconds': duracion_segundos,
                'ocurrencias': 1,
                'episodios': {episodio_actual}
            }

    stats['unique_tracks_count'] = len(pistas_agrupadas_temp)
    stats['pistas_consolidadas'] = sorted(pistas_agrupadas_temp.values(), key=lambda x: x['duration_seconds'], reverse=True)

    pistas_detalladas_temp = []
    for datos in pistas_agrupadas_temp.values():
        try:
             episodios_sorted_list = sorted(list(datos['episodios']), key=lambda ep: int(re.search(r'\d+', ep).group()) if re.search(r'\d+', ep) else float('inf'))
        except:
            episodios_sorted_list = sorted(list(datos['episodios']))
        pistas_detalladas_temp.append({
            'title': datos['title'], 'composer': datos['composer'], 'publisher': datos['publisher'],
            'count': datos['ocurrencias'], 'tiempo_total': datos['duration_seconds'],
            'tiempo_formateado': formatear_tiempo(datos['duration_seconds']),
            'episodios_count': len(datos['episodios']), 'episodios_lista': ', '.join(episodios_sorted_list)
        })
    stats['pistas_repetidas_detalle'] = sorted(pistas_detalladas_temp, key=lambda x: x['count'], reverse=True)

    def get_start_seconds(rango_str):
        try:
            minutes, seconds = map(int, rango_str.split('-')[0].split(':'))
            return minutes * 60 + seconds
        except: return float('inf')
    stats['pistas_por_duracion'] = Counter(dict(sorted(stats['pistas_por_duracion'].items(), key=lambda item: get_start_seconds(item[0]))))

    return stats


# =========================
# Funciones de Generación de Gráficos (Usa paleta actualizada v1.8.0)
# =========================
def generar_grafica_publishers(estadisticas: Dict[str, Any], output_dir: Path) -> Optional[str]:
    """ Genera gráfico circular de publishers (rosado/púrpura). Devuelve ruta ABSOLUTA o None. """
    if not MATPLOTLIB_AVAILABLE: return None
    publishers_tiempo_valid = {k: v for k, v in estadisticas.get('publishers_tiempo', Counter()).items() if k != 'N/A' and v > 0}
    if not publishers_tiempo_valid:
        print("ℹ️ Gráfico Publishers: No hay datos válidos.")
        return None

    total_tiempo = sum(publishers_tiempo_valid.values())
    if total_tiempo == 0: return None

    labels, sizes = [], []
    otros_tiempo, otros_count = 0, 0
    sorted_publishers = sorted(publishers_tiempo_valid.items(), key=lambda item: item[1], reverse=True)

    threshold_count = 5
    for name, tiempo in sorted_publishers:
        percentage = (tiempo / total_tiempo) * 100
        if percentage < PIE_CHART_OTHERS_THRESHOLD and len(sorted_publishers) > threshold_count:
            otros_tiempo += tiempo
            otros_count += 1
        else:
            label_name = (name[:30] + '...') if len(name) > 33 else name
            labels.append(f"{label_name}\n({formatear_tiempo(tiempo)})")
            sizes.append(tiempo)

    if otros_tiempo > 0:
        labels.append(f"Otros ({otros_count})\n({formatear_tiempo(otros_tiempo)})")
        sizes.append(otros_tiempo)

    if not labels: return None

    num_slices = len(sizes)

    try:
        fig, ax = plt.subplots(figsize=(12, 8))

        base_pie_colors = CHART_COLORS['pie'] # Usa la paleta definida
        if num_slices <= len(base_pie_colors):
            final_pie_colors = [base_pie_colors[i % len(base_pie_colors)] for i in range(num_slices)]
        else:
            try:
                colormap = plt.cm.get_cmap('tab20', num_slices) # Colormap como fallback
                final_pie_colors = [colormap(i) for i in range(num_slices)]
                print(f"INFO Gráfico Publishers: Usando colormap 'tab20' para {num_slices} slices.")
            except Exception as e_cmap:
                print(f"ADVERTENCIA Gráfico Publishers: Falló al usar colormap. Usando paleta base repetida. Error: {e_cmap}", file=sys.stderr)
                final_pie_colors = [base_pie_colors[i % len(base_pie_colors)] for i in range(num_slices)]

        wedges, texts, autotexts = ax.pie(
            sizes, autopct='%1.1f%%', startangle=90, pctdistance=0.85,
            colors=final_pie_colors, wedgeprops={'edgecolor': 'white', 'linewidth': 0.5}
        )
        ax.axis('equal')
        # <<< CAMBIO v1.8.0: Color de texto en % ajustado (era blanco) -> oscuro para mejor contraste con rosado/púrpura? >>>
        # Mantener blanco por ahora, suele verse bien en colores saturados.
        plt.setp(autotexts, size=8, weight="bold", color=COLOR_PALETTE['text_on_dark'])
        plt.title('Distribución de Tiempo por Editora (Publisher)', fontsize=16, pad=20, color=CHART_COLORS['titles'])

        legend = ax.legend(wedges, labels, title="Editoras", loc="center left", bbox_to_anchor=(1.05, 0, 0.5, 1), fontsize='small', frameon=False)
        plt.setp(legend.get_texts(), color=CHART_COLORS['axis_labels'])
        plt.setp(legend.get_title(), color=CHART_COLORS['titles'], weight='bold')

        plt.subplots_adjust(left=0.1, right=0.7, top=0.9, bottom=0.1)

        chart_path = output_dir / "publisher_pie_chart.png"
        plt.savefig(chart_path, bbox_inches='tight', dpi=150)
        plt.close(fig)
        print(f"✅ Gráfico de Publishers generado: {chart_path}")
        return str(chart_path.resolve())

    except Exception as e:
        print(f"❌ Error generando gráfico Publishers: {e}", file=sys.stderr)
        if 'fig' in locals() and plt.fignum_exists(fig.number): plt.close(fig)
        return None

def generar_grafica_compositores(estadisticas: Dict[str, Any], output_dir: Path) -> Optional[str]:
    """ Genera gráfico de barras de compositores (barras rosadas). Devuelve ruta ABSOLUTA o None. """
    if not MATPLOTLIB_AVAILABLE: return None
    top_compositores = sorted(
        [(c, t) for c, t in estadisticas.get('compositores_tiempo', Counter()).items() if c != 'N/A' and t > 0],
        key=lambda item: item[1],
        reverse=True
    )[:BAR_CHART_TOP_N_COMPOSERS]

    if not top_compositores:
        print("ℹ️ Gráfico Compositores: No hay datos válidos.")
        return None

    nombres = [item[0] for item in top_compositores]
    tiempos_sec = [item[1] for item in top_compositores]
    nombres.reverse()
    tiempos_sec.reverse()

    try:
        fig_height = max(6, len(nombres) * 0.45)
        fig, ax = plt.subplots(figsize=(10, fig_height))
        y_pos = np.arange(len(nombres))

        # <<< CAMBIO v1.8.0: Usa color definido en CHART_COLORS (rosado) >>>
        bars = ax.barh(y_pos, tiempos_sec, align='center', color=CHART_COLORS['bar_composers'], height=0.6)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(nombres, fontsize=9, color=CHART_COLORS['axis_labels'])
        ax.set_xlabel('Tiempo Total (MM:SS)', color=CHART_COLORS['axis_labels'])
        ax.set_title(f'Top {len(nombres)} Compositores por Tiempo Total', color=CHART_COLORS['titles'], fontsize=14, pad=15)

        ax.xaxis.set_major_formatter(mticker.FuncFormatter(time_formatter))
        ax.tick_params(axis='x', rotation=30, colors=CHART_COLORS['axis_labels'])
        ax.tick_params(axis='y', colors=CHART_COLORS['axis_labels'])

        max_time_val = max(tiempos_sec) if tiempos_sec else 1
        label_color = COLOR_PALETTE['text_on_light'] # Etiquetas de valor en oscuro
        for bar in bars:
            width = bar.get_width()
            label_text = formatear_tiempo(int(width))
            ax.text(width + max_time_val * 0.01, bar.get_y() + bar.get_height() / 2.,
                    label_text, va='center', ha='left', fontsize=8, color=label_color)

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(CHART_COLORS['grid'])
        ax.spines['bottom'].set_color(CHART_COLORS['grid'])
        ax.xaxis.grid(True, linestyle='--', alpha=0.6, color=CHART_COLORS['grid'])
        ax.set_axisbelow(True)

        plt.subplots_adjust(left=0.35, right=0.95, top=0.9, bottom=0.15)
        chart_path = output_dir / "composers_bar_chart.png"
        plt.savefig(chart_path, dpi=150)
        plt.close(fig)
        print(f"✅ Gráfico de Compositores generado: {chart_path}")
        return str(chart_path.resolve())

    except Exception as e:
        print(f"❌ Error generando gráfico Compositores: {e}", file=sys.stderr)
        if 'fig' in locals() and plt.fignum_exists(fig.number): plt.close(fig)
        return None

def generar_grafica_pistas_top_tiempo(estadisticas: Dict[str, Any], output_dir: Path) -> Optional[str]:
    """ Genera gráfico de barras de pistas por tiempo (barras púrpuras). Devuelve ruta ABSOLUTA o None. """
    if not MATPLOTLIB_AVAILABLE: return None
    top_pistas = estadisticas.get('pistas_consolidadas', [])[:BAR_CHART_TOP_N_TRACKS_TIME]

    if not top_pistas:
        print("ℹ️ Gráfico Pistas por Tiempo: No hay datos.")
        return None

    titulos_compositor = [f"{(p['title'][:40]+'...' if len(p['title'])>43 else p['title'])}\n({(p['composer'][:35]+'...' if len(p['composer'])>38 else p['composer'])})" for p in top_pistas]
    tiempos_sec = [p['duration_seconds'] for p in top_pistas]
    ocurrencias = [p['ocurrencias'] for p in top_pistas]
    titulos_compositor.reverse()
    tiempos_sec.reverse()
    ocurrencias.reverse()

    try:
        fig_height = max(6, len(titulos_compositor) * 0.55)
        fig, ax = plt.subplots(figsize=(12, fig_height))
        y_pos = np.arange(len(titulos_compositor))

        # <<< CAMBIO v1.8.0: Usa color definido en CHART_COLORS (púrpura) >>>
        bars = ax.barh(y_pos, tiempos_sec, align='center', color=CHART_COLORS['bar_tracks'], height=0.6)
        ax.set_yticks(y_pos)
        ytick_fontsize = 9 if len(titulos_compositor) < 15 else 8
        ax.set_yticklabels(titulos_compositor, fontsize=ytick_fontsize, color=CHART_COLORS['axis_labels'])
        ax.set_xlabel('Tiempo Total Acumulado (MM:SS)', color=CHART_COLORS['axis_labels'])
        ax.set_title(f'Top {len(titulos_compositor)} Pistas por Tiempo Total Acumulado', color=CHART_COLORS['titles'], fontsize=14, pad=15)

        ax.xaxis.set_major_formatter(mticker.FuncFormatter(time_formatter))
        ax.tick_params(axis='x', rotation=30, colors=CHART_COLORS['axis_labels'])
        ax.tick_params(axis='y', colors=CHART_COLORS['axis_labels'])

        max_time_val = max(tiempos_sec) if tiempos_sec else 1
        label_color = COLOR_PALETTE['text_on_light'] # Etiquetas de valor en oscuro
        for i, bar in enumerate(bars):
            width = bar.get_width()
            label_text = f"{formatear_tiempo(int(width))} ({ocurrencias[i]}x)"
            ax.text(width + max_time_val * 0.01, bar.get_y() + bar.get_height() / 2.,
                    label_text, va='center', ha='left', fontsize=8, color=label_color)

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color(CHART_COLORS['grid'])
        ax.spines['bottom'].set_color(CHART_COLORS['grid'])
        ax.xaxis.grid(True, linestyle='--', alpha=0.6, color=CHART_COLORS['grid'])
        ax.set_axisbelow(True)

        plt.subplots_adjust(left=0.4, right=0.95, top=0.9, bottom=0.15)
        chart_path = output_dir / "tracks_time_bar_chart.png"
        plt.savefig(chart_path, dpi=150)
        plt.close(fig)
        print(f"✅ Gráfico de Pistas por Tiempo generado: {chart_path}")
        return str(chart_path.resolve())

    except Exception as e:
        print(f"❌ Error generando gráfico Pistas por Tiempo: {e}", file=sys.stderr)
        if 'fig' in locals() and plt.fignum_exists(fig.number): plt.close(fig)
        return None

def generar_grafica_episodios(stats_por_episodio: Dict[str, Dict[str, Any]], output_dir: Path) -> Optional[str]:
    """ Genera gráfico comparativo por episodio (Minutos rosado vs Pistas Únicas púrpura). Devuelve ruta ABSOLUTA o None. """
    if not MATPLOTLIB_AVAILABLE or not stats_por_episodio:
        print("ℹ️ Gráfico Comparativo Episodios: No disponible.")
        return None

    try:
        episodios_ordenados = sorted(stats_por_episodio.keys(), key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf'))
    except Exception:
        episodios_ordenados = sorted(stats_por_episodio.keys())

    if not episodios_ordenados:
        print("ℹ️ Gráfico Comparativo Episodios: No hay episodios con estadísticas.")
        return None

    minutos_totales = [stats_por_episodio[ep].get('duracion_total_segundos', 0) / 60.0 for ep in episodios_ordenados]
    pistas_unicas = [stats_por_episodio[ep].get('unique_tracks', 0) for ep in episodios_ordenados]

    try:
        x = np.arange(len(episodios_ordenados))
        width = 0.35
        fig_width = max(10, len(episodios_ordenados) * 0.7)
        fig, ax1 = plt.subplots(figsize=(fig_width, 6))

        # Eje Y Izquierdo (Minutos) - Rosado
        # <<< CAMBIO v1.8.0: Usa colores definidos (rosado/púrpura) >>>
        color1 = CHART_COLORS['bar_episodes_minutes']
        ax1.set_xlabel('Episodio', color=CHART_COLORS['axis_labels'])
        ax1.set_ylabel('Minutos Totales de Música', color=color1)
        bars1 = ax1.bar(x - width/2, minutos_totales, width, label='Minutos Totales', color=color1)
        ax1.tick_params(axis='y', labelcolor=color1)
        ax1.tick_params(axis='x', colors=CHART_COLORS['axis_labels'], rotation=90)
        ax1.set_xticks(x)
        ax1.set_xticklabels(episodios_ordenados, fontsize=8)
        ax1.set_xlim(-0.5, len(episodios_ordenados) - 0.5)
        ax1.set_ylim(bottom=0)

        # Eje Y Derecho (Pistas Únicas) - Púrpura
        ax2 = ax1.twinx()
        color2 = CHART_COLORS['bar_episodes_tracks']
        ax2.set_ylabel('Nº Pistas Únicas', color=color2)
        bars2 = ax2.bar(x + width/2, pistas_unicas, width, label='Pistas Únicas', color=color2)
        ax2.tick_params(axis='y', labelcolor=color2)
        ax2.set_ylim(bottom=0)
        ax2.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))

        plt.title('Comparativa por Episodio: Minutos vs. Pistas Únicas', color=CHART_COLORS['titles'], fontsize=14, pad=15)

        ax1.spines['top'].set_visible(False)
        ax2.spines['top'].set_visible(False)
        ax1.spines['left'].set_color(color1)
        ax2.spines['right'].set_color(color2)
        ax1.spines['bottom'].set_color(CHART_COLORS['grid'])
        ax1.grid(True, axis='y', linestyle='--', alpha=0.6, color=CHART_COLORS['grid'])
        ax1.set_axisbelow(True)

        fig.tight_layout()
        chart_path = output_dir / "episodes_comparison_chart.png"
        plt.savefig(chart_path, dpi=150)
        plt.close(fig)
        print(f"✅ Gráfico Comparativo Episodios generado: {chart_path}")
        return str(chart_path.resolve())

    except Exception as e:
        print(f"❌ Error generando gráfico Comparativo Episodios: {e}", file=sys.stderr)
        if 'fig' in locals() and plt.fignum_exists(fig.number): plt.close(fig)
        return None


# =========================
# Funciones de Generación de Reportes
# =========================

# --- Reporte Individual (MD) ---
def generar_reporte_estadisticas(datos_tabla: List[Dict[str, Any]], episodio: str, output_dir: str) -> Optional[str]:
    """ Genera reporte Markdown de estadísticas para un episodio individual. """
    if not datos_tabla: return None
    print(f"Generando reporte de estadísticas para episodio {episodio} en '{output_dir}'...")
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    try:
        estadisticas = calcular_estadisticas(datos_tabla)
    except Exception as e:
        print(f"❌ Error calculando estadísticas para Ep {episodio}: {e}", file=sys.stderr)
        return None

    total_pistas = estadisticas['total_pistas']
    segundos_totales = estadisticas['duracion_total_segundos']
    tiempo_formateado_total = formatear_tiempo(segundos_totales)
    promedio_duracion_seg = segundos_totales // total_pistas if total_pistas > 0 else 0
    tiempo_promedio = formatear_tiempo(promedio_duracion_seg)
    compositores_unicos = len(estadisticas['compositores'])
    publishers_unicos = len(estadisticas['publishers'])
    pistas_unicas_epi = estadisticas.get('unique_tracks_count', 0)

    # Construir contenido Markdown
    contenido = f"# Reporte de Música - Episodio {episodio}\n\n## Resumen Ejecutivo\n\n"
    contenido += "| Métrica                    | Valor          |\n|:---------------------------|---------------:|\n"
    contenido += f"| Total de pistas (usos)     | {total_pistas} |\n"
    contenido += f"| Pistas Únicas (Título+Comp)| {pistas_unicas_epi} |\n"
    contenido += f"| Tiempo total de música     | {tiempo_formateado_total} |\n"
    contenido += f"| Duración promedio / pista  | {tiempo_promedio} |\n"
    contenido += f"| Compositores únicos        | {compositores_unicos} |\n"
    contenido += f"| Editoras únicas (Publishers)| {publishers_unicos} |\n\n"

    pistas_consolidadas = estadisticas.get('pistas_consolidadas', [])
    if pistas_consolidadas:
        top_n = min(TOP_N_PISTAS_EPISODIO, len(pistas_consolidadas))
        contenido += f"## Top {top_n} Pistas (por Tiempo Total Acumulado en Episodio)\n\n"
        contenido += "| # | Título | Compositor | Editora | Tiempo Total | Ocurrencias |\n|:--|:-------|:-----------|:----------|:------------:|:-----------:|\n"
        for i, p in enumerate(pistas_consolidadas[:top_n], 1):
            tiempo_pista_fmt = formatear_tiempo(p['duration_seconds'])
            titulo_corto = (p['title'][:30]+'...') if len(p['title'])>33 else p['title']
            compositor_corto = (p['composer'][:25]+'...') if len(p['composer'])>28 else p['composer']
            publisher_corto = (p['publisher'][:25]+'...') if len(p['publisher'])>28 else p['publisher']
            contenido += f"| {i} | {titulo_corto} | {compositor_corto} | {publisher_corto} | {tiempo_pista_fmt} | {p['ocurrencias']} |\n"
        contenido += "\n"

    compositores_por_tiempo = sorted(estadisticas['compositores_tiempo'].items(), key=lambda item: item[1], reverse=True)
    compositores_validos = [(c, t) for c, t in compositores_por_tiempo if c != 'N/A' and t > 0]
    if compositores_validos:
        top_n = min(TOP_N_COMPOSITOR_EPISODIO, len(compositores_validos))
        contenido += f"## Top {top_n} Compositores (por Tiempo Total en Episodio)\n\n"
        contenido += "| Compositor | Pistas | Tiempo Total |\n|:-----------|:------:|:------------:|\n"
        for c, t_sec in compositores_validos[:top_n]:
            count = estadisticas['compositores'].get(c, 0)
            contenido += f"| {c} | {count} | {formatear_tiempo(t_sec)} |\n"
        contenido += "\n"

    publishers_por_tiempo = sorted(estadisticas['publishers_tiempo'].items(), key=lambda item: item[1], reverse=True)
    publishers_validos = [(p, t) for p, t in publishers_por_tiempo if p != 'N/A' and t > 0]
    if publishers_validos:
        top_n = min(TOP_N_PUBLISHER_EPISODIO, len(publishers_validos))
        contenido += f"## Top {top_n} Editoras (Publishers) (por Tiempo Total en Episodio)\n\n"
        contenido += "| Editora | Pistas | Tiempo Total |\n|:----------|:------:|:------------:|\n"
        for p, t_sec in publishers_validos[:top_n]:
            count = estadisticas['publishers'].get(p, 0)
            contenido += f"| {p} | {count} | {formatear_tiempo(t_sec)} |\n"
        contenido += "\n"

    pistas_por_duracion_ordenado = estadisticas.get('pistas_por_duracion', {}).items()
    if pistas_por_duracion_ordenado:
        contenido += "## Distribución por Duración de Pista\n\n"
        contenido += "| Rango (MM:SS)  | Número de Pistas | % del Total |\n|:---------------|:----------------:|:-----------:|\n"
        for rango, count in pistas_por_duracion_ordenado:
            porcentaje = (count / total_pistas) * 100 if total_pistas > 0 else 0
            contenido += f"| {rango} | {count} | {porcentaje:.1f}% |\n"
        contenido += "\n"

    contenido += "## Resumen del Episodio (Puntos Clave)\n\n"
    top_c = compositores_validos[0][0] if compositores_validos else "N/A"
    top_p = publishers_validos[0][0] if publishers_validos else "N/A"
    top_pista_info = pistas_consolidadas[0] if pistas_consolidadas else None
    top_pista_titulo = top_pista_info['title'] if top_pista_info else "N/A"
    contenido += f"- **Compositor principal (por tiempo):** {top_c}\n"
    contenido += f"- **Editora principal (por tiempo):** {top_p}\n"
    contenido += f"- **Pista principal (por tiempo acumulado):** {top_pista_titulo}\n"

    nombre_archivo_reporte = REPORTE_EPISODIO_FILENAME_FORMAT.format(episodio=episodio)
    report_file_path = output_path / nombre_archivo_reporte
    try:
        with open(report_file_path, 'w', encoding='utf-8') as f: f.write(contenido)
        print(f"✅ Reporte de estadísticas (MD) generado para Ep {episodio}: {report_file_path}")
        return str(report_file_path)
    except IOError as e:
        print(f"❌ Error al escribir el reporte MD para Ep {episodio} en {report_file_path}: {e}", file=sys.stderr)
        return None


# --- Reporte Global PDF ---
# (Usa la clase PDFReport actualizada con colores v1.8.0)
def generar_reporte_global_pdf(datos_consolidados: List[Dict[str, Any]],
                               estadisticas: Dict[str, Any],
                               stats_por_episodio: Dict[str, Dict[str, Any]],
                               episodios_ordenados: List[str],
                               chart_paths: Dict[str, Optional[str]],
                               output_dir_path: Path,
                               report_name: str) -> Optional[str]:
    """ Genera el reporte global en formato PDF con diseño y nombre personalizado. """
    if not FPDF2_AVAILABLE: return None
    if not datos_consolidados: return None

    print(f"Generando reporte global PDF '{report_name}_Global.pdf' en '{output_dir_path}'...")

    try:
        pdf = PDFReport(report_name=report_name, orientation='P', unit='mm', format='A4')
        pdf.chart_paths = {k: v for k, v in chart_paths.items() if v}
        pdf.add_page()

        # Título Principal y Episodios (Page 1)
        # <<< CAMBIO v1.8.0: Usar report_name y título genérico >>>
        pdf.chapter_title(f"{report_name}\nReporte Global de Música", level=1)
        if episodios_ordenados:
            pdf.set_font(pdf.DEFAULT_FONT, '', pdf.FONT_SIZE_BODY - 1)
            pdf.multi_cell(0, pdf.font_size * pdf.LINE_HEIGHT_MULTIPLIER, f"Episodios incluidos ({len(episodios_ordenados)}): {', '.join(episodios_ordenados)}", border=0, align='L', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(pdf.SPACING_AFTER_TEXT * 2)
        else:
            pdf.body_text("*No se procesaron episodios o no se encontraron datos.*")
            pdf.ln(pdf.SPACING_AFTER_TEXT * 2)

        # Resumen General (Page 1 - Con colores actualizados v1.8.0)
        total_pistas = estadisticas['total_pistas']
        segundos_totales = estadisticas['duracion_total_segundos']
        tiempo_formateado_total = formatear_tiempo(segundos_totales)
        promedio_duracion_seg = segundos_totales // total_pistas if total_pistas > 0 else 0
        tiempo_promedio = formatear_tiempo(promedio_duracion_seg)
        pistas_unicas_global = estadisticas.get('unique_tracks_count', 0)
        compositores_unicos = len(estadisticas.get('compositores', {}))
        publishers_unicos = len(estadisticas.get('publishers', {}))

        rhapsody_time_sec = estadisticas.get('publishers_tiempo', {}).get(PUBLISHER_RHAPSODY, 0)
        rhapsody_time_fmt = formatear_tiempo(rhapsody_time_sec)
        rhapsody_perc = (rhapsody_time_sec / segundos_totales * 100) if segundos_totales > 0 else 0
        rhapsody_composers_time = Counter()
        for pista in datos_consolidados:
            if PUBLISHER_RHAPSODY in pista.get('publisher', ''):
                 composers = pista.get('composer', 'N/A').split(' / ')
                 duration = pista.get('duration_seconds', 0)
                 for comp in composers:
                     comp_limpio = comp.strip()
                     if comp_limpio and comp_limpio != 'N/A': rhapsody_composers_time[comp_limpio] += duration
        sorted_rhapsody_composers = sorted(rhapsody_composers_time.items(), key=lambda item: item[1], reverse=True)

        resumen_data_main = [
            ["Total de episodios", str(len(episodios_ordenados))],
            ["Total de pistas (usos)", f"{total_pistas:,}".replace(",", ".")],
            ["Pistas Únicas (Título+Comp)", f"{pistas_unicas_global:,}".replace(",", ".")],
            ["Tiempo total de música (MM:SS)", tiempo_formateado_total],
            ["Duración promedio / pista (MM:SS)", tiempo_promedio],
            ["Compositores únicos", str(compositores_unicos)],
            ["Editoras únicas", str(publishers_unicos)]
        ]
        resumen_data_extra = [
             ["Tiempo Total Rhapsolody", rhapsody_time_fmt],
             ["% Tiempo Rhapsolody", f"{rhapsody_perc:.1f}%"]
        ]
        num_comp_to_show = 5
        for i, (comp, time_sec) in enumerate(sorted_rhapsody_composers[:num_comp_to_show]):
             resumen_data_extra.append([f"  └ {comp}", formatear_tiempo(time_sec)])
        if len(sorted_rhapsody_composers) > num_comp_to_show:
             resumen_data_extra.append([f"  └ Otros ({len(sorted_rhapsody_composers) - num_comp_to_show})...", ""])

        # Dibujar el resumen (usará colores actualizados internamente)
        pdf.add_resumen_general(data=resumen_data_main, title="Resumen General", extra_data=resumen_data_extra)

        # Forzar salto de página si aún estamos en la página 1
        if pdf.page_no() == 1: pdf.add_page()

        # === INICIO SECCIONES CON SALTOS DE PÁGINA ===

        # Sección: Análisis por Editora
        publishers_por_tiempo = sorted([(p, estadisticas['publishers'][p], t) for p, t in estadisticas.get('publishers_tiempo', {}).items() if p != 'N/A' and t > 0], key=lambda x: x[2], reverse=True)
        if publishers_por_tiempo:
            pdf.chapter_title("Análisis por Editora (Publisher)", level=2) # Título H2 púrpura
            pdf.add_chart('publishers', "Distribución de Tiempo por Editora")
            top_n_publishers = min(TOP_N_PUBLISHER_GLOBAL, len(publishers_por_tiempo))
            pub_headers = ["Editora", "Pistas", "T. Total"]
            pub_data = [[pub, str(count), formatear_tiempo(t_sec)] for pub, count, t_sec in publishers_por_tiempo[:top_n_publishers]]
            pub_col_widths = [120, 25, 35]
            pdf.add_table(headers=pub_headers, data=pub_data, col_widths=pub_col_widths, title=f"Tabla: Top {top_n_publishers} Editoras por Tiempo")

            top_publisher_nombre = publishers_por_tiempo[0][0]
            pistas_repetidas_lista = estadisticas.get('pistas_repetidas_detalle', [])
            pistas_top_pub = sorted([p for p in pistas_repetidas_lista if top_publisher_nombre in p['publisher'].split(' / ')], key=lambda x: x['tiempo_total'], reverse=True)
            top_n_detail_pub = min(TOP_N_PISTAS_DETALLE_GLOBAL, len(pistas_top_pub))
            if top_n_detail_pub > 0:
                det_pub_headers = ["Título", "Compositor", "Reps", "T.Total"]
                det_pub_data = []
                for pista in pistas_top_pub[:top_n_detail_pub]:
                    titulo_det = (pista['title'][:65] + '...') if len(pista['title']) > 68 else pista['title']
                    comp_det = (pista['composer'][:55] + '...') if len(pista['composer']) > 58 else pista['composer']
                    det_pub_data.append([titulo_det, comp_det, str(pista['count']), pista['tiempo_formateado']])
                det_pub_col_widths = [85, 65, 12, 18]
                pdf.add_table(headers=det_pub_headers, data=det_pub_data, col_widths=det_pub_col_widths, title=f"Top {top_n_detail_pub} Pistas de {top_publisher_nombre}")

        # Sección: Resumen por Episodio
        if stats_por_episodio and episodios_ordenados:
            pdf.add_page()
            pdf.chapter_title("Resumen y Comparativa por Episodio", level=2) # Título H2 púrpura
            pdf.add_chart('episodes_comparison', "Comparativa Gráfica: Minutos vs. Pistas Únicas")
            ep_headers = ["Episodio", "Pistas", "P. Únicas", "Duración (MM:SS)"]
            ep_data = []
            for ep_id in episodios_ordenados:
                s = stats_por_episodio.get(ep_id, {})
                ep_data.append([
                    ep_id, str(s.get('pistas', 0)), str(s.get('unique_tracks', 0)), s.get('duracion_formateada', '00:00')
                ])
            ep_col_widths = [30, 30, 30, 90]
            pdf.add_table(headers=ep_headers, data=ep_data, col_widths=ep_col_widths, title="Tabla de Resumen por Episodio")

        # Sección: Análisis de Tendencias (Pistas)
        pistas_repetidas_lista = estadisticas.get('pistas_repetidas_detalle', [])
        if pistas_repetidas_lista:
            pdf.add_page()
            pdf.chapter_title("Análisis de Tendencias (Pistas)", level=2) # Título H2 púrpura
            top_n_ocurrencias = min(TOP_N_PISTAS_GLOBAL, len(pistas_repetidas_lista))
            trend_headers = ["#", "Título", "Compositor", "Eps", "Reps", "T.Total"]
            trend_data = []
            for i, p in enumerate(pistas_repetidas_lista[:top_n_ocurrencias], 1):
                titulo_pdf = (p['title'][:60] + '...') if len(p['title']) > 63 else p['title']
                comp_pdf = (p['composer'][:45] + '...') if len(p['composer']) > 48 else p['composer']
                trend_data.append([ str(i), titulo_pdf, comp_pdf, str(p['episodios_count']), str(p['count']), p['tiempo_formateado'] ])
            trend_col_widths = [8, 75, 55, 12, 12, 18]
            pdf.add_table(headers=trend_headers, data=trend_data, col_widths=trend_col_widths, title=f"Top {top_n_ocurrencias} Pistas por Ocurrencias")
            pdf.add_chart('tracks_time', f"Top {BAR_CHART_TOP_N_TRACKS_TIME} Pistas por Tiempo Total Acumulado")

        # Sección: Análisis por Compositor
        compositores_por_tiempo = sorted([(c, estadisticas['compositores'][c], t) for c, t in estadisticas.get('compositores_tiempo', {}).items() if c != 'N/A' and t > 0], key=lambda x: x[2], reverse=True)
        if compositores_por_tiempo:
            pdf.add_page()
            pdf.chapter_title("Análisis por Compositor", level=2) # Título H2 púrpura
            pdf.add_chart('composers', f"Top {BAR_CHART_TOP_N_COMPOSERS} Compositores por Tiempo Total")
            top_n_compositores = min(TOP_N_COMPOSITOR_GLOBAL, len(compositores_por_tiempo))
            comp_headers = ["Compositor", "Pistas", "T. Total"]
            comp_data = [[c, str(count), formatear_tiempo(t_sec)] for c, count, t_sec in compositores_por_tiempo[:top_n_compositores]]
            comp_col_widths = [120, 25, 35]
            pdf.add_table(headers=comp_headers, data=comp_data, col_widths=comp_col_widths, title=f"Tabla: Top {top_n_compositores} Compositores por Tiempo")

            top_compositor_nombre = compositores_por_tiempo[0][0]
            pistas_top_comp = sorted([p for p in pistas_repetidas_lista if top_compositor_nombre in p['composer'].split(' / ')], key=lambda x: x['tiempo_total'], reverse=True)
            top_n_detail_comp = min(TOP_N_PISTAS_DETALLE_GLOBAL, len(pistas_top_comp))
            if top_n_detail_comp > 0:
                det_comp_headers = ["Título", "Editora", "Reps", "T.Total"]
                det_comp_data = []
                for pista in pistas_top_comp[:top_n_detail_comp]:
                    titulo_det = (pista['title'][:65] + '...') if len(pista['title']) > 68 else pista['title']
                    pub_det = (pista['publisher'][:55] + '...') if len(pista['publisher']) > 58 else pista['publisher']
                    det_comp_data.append([titulo_det, pub_det, str(pista['count']), pista['tiempo_formateado']])
                det_comp_col_widths = [85, 65, 12, 18]
                pdf.add_table(headers=det_comp_headers, data=det_comp_data, col_widths=det_comp_col_widths, title=f"Top {top_n_detail_comp} Pistas de {top_compositor_nombre}")

        # Sección: Distribución por Duración
        pistas_por_duracion_ordenado = estadisticas.get('pistas_por_duracion', {}).items()
        if pistas_por_duracion_ordenado:
            pdf.add_page()
            pdf.chapter_title("Distribución por Duración de Pista", level=2) # Título H2 púrpura
            dist_headers = ["Rango (MM:SS)", "Nº Pistas", "% del Total"]
            dist_data = []
            total_pistas_dist = sum(count for _, count in pistas_por_duracion_ordenado)
            for rango, count in pistas_por_duracion_ordenado:
                porcentaje = (count / total_pistas_dist) * 100 if total_pistas_dist > 0 else 0
                dist_data.append([rango, f"{count:,}".replace(",", "."), f"{porcentaje:.1f}%"])
            dist_col_widths = [60, 60, 60]
            pdf.add_table(headers=dist_headers, data=dist_data, col_widths=dist_col_widths)

        # === FIN SECCIONES CON SALTOS DE PÁGINA ===

        pdf_filename = REPORTE_GLOBAL_PDF_FILENAME_FORMAT.format(report_name=report_name)
        pdf_file_path = output_dir_path / pdf_filename
        pdf.output(pdf_file_path)
        print(f"✅ Reporte global PDF generado con éxito: {pdf_file_path}")
        return str(pdf_file_path)

    except Exception as e:
        print(f"❌ Error crítico al generar el reporte PDF {report_name}_Global.pdf: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return None


# --- Reporte Global Markdown ---
def generar_reporte_global_md(datos_consolidados: List[Dict[str, Any]],
                              estadisticas: Dict[str, Any],
                              stats_por_episodio: Dict[str, Dict[str, Any]],
                              episodios_ordenados: List[str],
                              chart_paths: Dict[str, Optional[str]],
                              output_dir_path: Path,
                              report_name: str) -> Optional[str]:
    """ Genera el reporte global en formato MARKDOWN con nombre personalizado. """
    if not datos_consolidados: return None

    print(f"Generando reporte global MD '{report_name}_Global.md' en '{output_dir_path}'...")

    stats = estadisticas
    total_pistas = stats.get('total_pistas', 0)
    segundos_totales = stats.get('duracion_total_segundos', 0)
    tiempo_formateado_total = formatear_tiempo(segundos_totales)
    promedio_duracion_seg = segundos_totales // total_pistas if total_pistas > 0 else 0
    tiempo_promedio = formatear_tiempo(promedio_duracion_seg)
    pistas_unicas_global = stats.get('unique_tracks_count', 0)
    compositores_unicos = len(stats.get('compositores', {}))
    publishers_unicos = len(stats.get('publishers', {}))

    chart_paths_rel = {}
    for key, abs_path_str in chart_paths.items():
        if abs_path_str:
            try:
                 abs_path = Path(abs_path_str)
                 relative_path = Path(CHARTS_SUBDIR) / abs_path.name
                 chart_paths_rel[key] = relative_path.as_posix()
            except Exception as e: chart_paths_rel[key] = None
        else: chart_paths_rel[key] = None

    # Construir Contenido Markdown
    contenido = f"# {report_name} - Reporte Global de Música\n\n" # Usa report_name
    if episodios_ordenados:
        contenido += f"*Episodios incluidos ({len(episodios_ordenados)}): {', '.join(episodios_ordenados)}*\n\n"
    else:
        contenido += "*No se procesaron episodios.*\n\n"

    contenido += "## Resumen General\n\n"
    contenido += "| Métrica                      | Valor             |\n|:-----------------------------|------------------:|\n"
    contenido += f"| Total de episodios           | {len(episodios_ordenados)} |\n"
    contenido += f"| Total de pistas (usos)       | {total_pistas:,} |\n"
    contenido += f"| Pistas Únicas (Título+Comp)  | {pistas_unicas_global:,} |\n"
    contenido += f"| Tiempo total música (MM:SS)  | {tiempo_formateado_total} |\n"
    contenido += f"| Duración promedio/pista (MM:SS)| {tiempo_promedio} |\n"
    contenido += f"| Compositores únicos          | {compositores_unicos} |\n"
    contenido += f"| Editoras únicas              | {publishers_unicos} |\n\n"

    if stats_por_episodio and episodios_ordenados:
        contenido += "## Resumen y Comparativa por Episodio\n\n"
        chart_path_ep = chart_paths_rel.get('episodes_comparison')
        if chart_path_ep: contenido += f"![Comparativa Gráfica Episodios]({chart_path_ep})\n\n"
        elif MATPLOTLIB_AVAILABLE: contenido += "*Nota: No se pudo generar gráfico comparativo.*\n\n"
        else: contenido += "*Nota: Gráfico no generado (matplotlib no disponible).*\n\n"
        contenido += "| Episodio | Pistas | P. Únicas | Duración (MM:SS) |\n|:---------|:------:|:---------:|:----------------:|\n"
        for ep_id in episodios_ordenados:
             s = stats_por_episodio.get(ep_id, {})
             contenido += f"| {ep_id} | {s.get('pistas', 0)} | {s.get('unique_tracks', 0)} | {s.get('duracion_formateada', '00:00')} |\n"
        contenido += "\n"

    pistas_repetidas_lista = stats.get('pistas_repetidas_detalle', [])
    if pistas_repetidas_lista:
        contenido += "## Análisis de Tendencias (Pistas)\n\n"
        top_n_ocurrencias = min(TOP_N_PISTAS_GLOBAL, len(pistas_repetidas_lista))
        contenido += f"### Top {top_n_ocurrencias} Pistas Más Utilizadas (por Ocurrencias)\n\n"
        contenido += "| # | Título | Compositor | Eps | Reps | T.Total (MM:SS) |\n|:--|:-------|:-----------|:---:|:----:|:---------------:|\n"
        for i, p in enumerate(pistas_repetidas_lista[:top_n_ocurrencias], 1):
            titulo_md = (p['title'][:40]+'...') if len(p['title'])>43 else p['title']
            comp_md = (p['composer'][:35]+'...') if len(p['composer'])>38 else p['composer']
            contenido += f"| {i} | {titulo_md} | {comp_md} | {p['episodios_count']} | {p['count']} | {p['tiempo_formateado']} |\n"
        contenido += "\n"

        chart_path_trk = chart_paths_rel.get('tracks_time')
        if chart_path_trk:
             contenido += f"### Top {BAR_CHART_TOP_N_TRACKS_TIME} Pistas por Tiempo Total Acumulado\n\n![Top Pistas por Tiempo]({chart_path_trk})\n\n"
        elif MATPLOTLIB_AVAILABLE: contenido += "*Nota: No se pudo generar gráfico.*\n\n"
        else: contenido += "*Nota: Gráfico no generado (matplotlib no disponible).*\n\n"

    compositores_por_tiempo = sorted([(c, stats['compositores'][c], t) for c, t in stats.get('compositores_tiempo', {}).items() if c != 'N/A' and t > 0], key=lambda x: x[2], reverse=True)
    if compositores_por_tiempo:
        contenido += "## Análisis por Compositor\n\n"
        chart_path_comp = chart_paths_rel.get('composers')
        if chart_path_comp:
            contenido += f"### Top {BAR_CHART_TOP_N_COMPOSERS} Compositores por Tiempo Total\n\n![Top Compositores por Tiempo]({chart_path_comp})\n\n"
        elif MATPLOTLIB_AVAILABLE: contenido += "*Nota: No se pudo generar gráfico.*\n\n"
        else: contenido += "*Nota: Gráfico no generado (matplotlib no disponible).*\n\n"
        top_n_compositores = min(TOP_N_COMPOSITOR_GLOBAL, len(compositores_por_tiempo))
        contenido += f"#### Tabla: Top {top_n_compositores} Compositores por Tiempo\n\n| Compositor | Pistas | Tiempo Total (MM:SS) |\n|:-----------|:------:|:--------------------:|\n"
        for c, count, t_sec in compositores_por_tiempo[:top_n_compositores]: contenido += f"| {c} | {count} | {formatear_tiempo(t_sec)} |\n"
        contenido += "\n"

    publishers_por_tiempo = sorted([(p, stats['publishers'][p], t) for p, t in stats.get('publishers_tiempo', {}).items() if p != 'N/A' and t > 0], key=lambda x: x[2], reverse=True)
    if publishers_por_tiempo:
        contenido += "## Análisis por Editora (Publisher)\n\n"
        chart_path_pub = chart_paths_rel.get('publishers')
        if chart_path_pub: contenido += f"### Distribución de Tiempo por Editora\n\n![Distribución de Tiempo por Editora]({chart_path_pub})\n\n"
        elif MATPLOTLIB_AVAILABLE: contenido += "*Nota: No se pudo generar gráfico.*\n\n"
        else: contenido += "*Nota: Gráfico no generado (matplotlib no disponible).*\n\n"
        top_n_publishers = min(TOP_N_PUBLISHER_GLOBAL, len(publishers_por_tiempo))
        contenido += f"#### Tabla: Top {top_n_publishers} Editoras por Tiempo\n\n| Editora | Pistas | Tiempo Total (MM:SS) |\n|:----------|:------:|:--------------------:|\n"
        for p, count, t_sec in publishers_por_tiempo[:top_n_publishers]: contenido += f"| {p} | {count} | {formatear_tiempo(t_sec)} |\n"
        contenido += "\n"

    show_detail_section_md = (compositores_por_tiempo or publishers_por_tiempo) and pistas_repetidas_lista
    if show_detail_section_md:
        contenido += "## Detalle Pistas Principales\n"
        if compositores_por_tiempo:
            top_comp_nombre = compositores_por_tiempo[0][0]
            pistas_top_comp = sorted([p for p in pistas_repetidas_lista if top_comp_nombre in p['composer'].split(' / ')], key=lambda x: x['tiempo_total'], reverse=True)
            top_n_comp = min(TOP_N_PISTAS_DETALLE_GLOBAL, len(pistas_top_comp))
            if top_n_comp > 0:
                contenido += f"\n### Top {top_n_comp} Pistas de {top_comp_nombre}\n\n| Título | Editora | Reps | T.Total (MM:SS) |\n|:-------|:----------|:----:|:---------------:|\n"
                for p in pistas_top_comp[:top_n_comp]:
                     titulo_md_det = (p['title'][:35]+'...' if len(p['title'])>38 else p['title'])
                     pub_md_det = (p['publisher'][:30]+'...' if len(p['publisher'])>33 else p['publisher'])
                     contenido += f"| {titulo_md_det} | {pub_md_det} | {p['count']} | {p['tiempo_formateado']} |\n"
                contenido += "\n"
        if publishers_por_tiempo:
            top_pub_nombre = publishers_por_tiempo[0][0]
            pistas_top_pub = sorted([p for p in pistas_repetidas_lista if top_pub_nombre in p['publisher'].split(' / ')], key=lambda x: x['tiempo_total'], reverse=True)
            top_n_pub = min(TOP_N_PISTAS_DETALLE_GLOBAL, len(pistas_top_pub))
            if top_n_pub > 0:
                contenido += f"\n### Top {top_n_pub} Pistas de {top_pub_nombre}\n\n| Título | Compositor | Reps | T.Total (MM:SS) |\n|:-------|:-----------|:----:|:---------------:|\n"
                for p in pistas_top_pub[:top_n_pub]:
                    titulo_md_det = (p['title'][:35]+'...' if len(p['title'])>38 else p['title'])
                    comp_md_det = (p['composer'][:30]+'...' if len(p['composer'])>33 else p['composer'])
                    contenido += f"| {titulo_md_det} | {comp_md_det} | {p['count']} | {p['tiempo_formateado']} |\n"
                contenido += "\n"

    pistas_por_duracion_ordenado = stats.get('pistas_por_duracion', {}).items()
    if pistas_por_duracion_ordenado:
        contenido += "## Distribución por Duración de Pista\n\n| Rango (MM:SS)  | Número de Pistas | % del Total |\n|:---------------|:----------------:|:-----------:|\n"
        total_pistas_dist = sum(count for _, count in pistas_por_duracion_ordenado)
        for rango, count in pistas_por_duracion_ordenado:
            porcentaje = (count / total_pistas_dist) * 100 if total_pistas_dist > 0 else 0
            contenido += f"| {rango} | {count:,} | {porcentaje:.1f}% |\n"
        contenido += "\n"

    md_filename = REPORTE_GLOBAL_MD_FILENAME_FORMAT.format(report_name=report_name)
    md_file_path = output_dir_path / md_filename
    try:
        with open(md_file_path, 'w', encoding='utf-8') as f: f.write(contenido)
        print(f"✅ Reporte global Markdown generado: {md_file_path}")
        return str(md_file_path)
    except IOError as e:
        print(f"❌ Error al escribir el reporte MD global {md_file_path}: {e}", file=sys.stderr)
        return None


# --- Orquestador Reportes Globales ---
def generar_reportes_globales(datos_consolidados: List[Dict[str, Any]], output_dir: str, report_name: str):
    """ Orquestador principal para reportes globales (MD y PDF), usando nombre personalizado. """
    if not datos_consolidados: return

    output_path = Path(output_dir)
    charts_dir = output_path / CHARTS_SUBDIR
    charts_dir.mkdir(parents=True, exist_ok=True)

    print("\n--- Calculando Estadísticas Globales ---")
    try:
        estadisticas_globales = calcular_estadisticas(datos_consolidados)
        print("✅ Estadísticas globales calculadas.")
    except Exception as e_stats:
        print(f"❌ Error crítico al calcular estadísticas globales: {e_stats}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr); return

    print("--- Calculando Estadísticas por Episodio ---")
    stats_por_episodio = {}
    episodios_ordenados = []
    try:
        def sort_key_episodio(ep_id):
            match = re.search(r'\d+', ep_id); return int(match.group()) if match else float('inf')
        episodios_presentes = sorted(list(set(p['episode'] for p in datos_consolidados)), key=sort_key_episodio)
        episodios_ordenados = episodios_presentes

        for episodio_id in episodios_ordenados:
            pistas_episodio = [p for p in datos_consolidados if p.get('episode') == episodio_id]
            if pistas_episodio:
                 total_pistas_epi = len(pistas_episodio)
                 duracion_seg_epi = sum(p.get('duration_seconds', 0) for p in pistas_episodio)
                 unique_tracks_epi_set = set(f"{p.get('title', 'N/A').strip()}|{p.get('composer', 'N/A')}" for p in pistas_episodio)
                 stats_por_episodio[episodio_id] = {
                     'pistas': total_pistas_epi, 'duracion_total_segundos': duracion_seg_epi,
                     'duracion_formateada': formatear_tiempo(duracion_seg_epi), 'unique_tracks': len(unique_tracks_epi_set)
                 }
            else:
                 stats_por_episodio[episodio_id] = {'pistas': 0, 'duracion_total_segundos': 0, 'duracion_formateada': '00:00', 'unique_tracks': 0}
        print(f"✅ Estadísticas calculadas para {len(stats_por_episodio)} episodios.")
    except Exception as e_stats_epi:
        print(f"❌ Error calculando estadísticas por episodio: {e_stats_epi}", file=sys.stderr)
        stats_por_episodio = {}
        episodios_ordenados = sorted(list(set(p.get('episode', DEFAULT_EPISODIO) for p in datos_consolidados)))

    print("\n--- Generando Gráficos Globales ---")
    chart_paths = {'publishers': None, 'composers': None, 'tracks_time': None, 'episodes_comparison': None}
    if MATPLOTLIB_AVAILABLE:
        chart_paths['publishers'] = generar_grafica_publishers(estadisticas_globales, charts_dir)
        chart_paths['composers'] = generar_grafica_compositores(estadisticas_globales, charts_dir)
        chart_paths['tracks_time'] = generar_grafica_pistas_top_tiempo(estadisticas_globales, charts_dir)
        if stats_por_episodio:
            chart_paths['episodes_comparison'] = generar_grafica_episodios(stats_por_episodio, charts_dir)
        else:
            print("ℹ️ Gráfico comparativo de episodios omitido.")
    else: print("INFO: Generación de gráficos omitida (matplotlib no disponible).")
    print("--- Fin Generación Gráficos ---\n")

    # Generar Reporte Global Markdown
    try:
        generar_reporte_global_md(datos_consolidados, estadisticas_globales, stats_por_episodio, episodios_ordenados, chart_paths, output_path, report_name)
    except Exception as e_md:
        print(f"❌ Error crítico al generar el reporte global Markdown: {e_md}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)

    # Generar Reporte Global PDF
    try:
        generar_reporte_global_pdf(datos_consolidados, estadisticas_globales, stats_por_episodio, episodios_ordenados, chart_paths, output_path, report_name)
    except Exception as e_pdf:
        print(f"❌ Error al llamar a la generación del PDF global.", file=sys.stderr)


# ============================================
# Funciones de Procesamiento de Archivo Único
# ============================================
def _procesar_excel_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, episodio: str) -> List[Dict[str, Any]]:
    """ Extrae datos de una hoja de cálculo Excel específica. """
    datos_tabla = []
    contador_seq = 1
    for i in range(ROW_START, ws.max_row + 1):
        try:
            titulo = ws.cell(row=i, column=COL_TITULO).value
            tiempo_excel = ws.cell(row=i, column=COL_TIEMPO).value
            compositor_val = ws.cell(row=i, column=COL_COMPOSITOR).value
            publisher_val = ws.cell(row=i, column=COL_PUBLISHER).value

            if not titulo or not isinstance(titulo, str) or not titulo.strip() or tiempo_excel is None:
                continue

            titulo = titulo.strip()
            tiempo_formateado, duracion_segundos = parsear_y_formatear_tiempo(tiempo_excel)
            if duracion_segundos <= 0: continue

            compositor_str = limpiar_participante(compositor_val)
            publisher_str = limpiar_participante(publisher_val)

            datos_tabla.append({
                'seq': contador_seq, 'title': titulo, 'publisher': publisher_str,
                'composer': compositor_str, 'time': tiempo_formateado,
                'duration_seconds': duracion_segundos, 'episode': episodio
            })
            contador_seq += 1
        except Exception as e_row:
            print(f"❌ Error procesando fila Excel {i} (Ep {episodio}): {e_row}", file=sys.stderr)
    return datos_tabla

def _procesar_markdown_sheet(archivo_md_path: Path, episodio: str) -> List[Dict[str, Any]]:
    """ Extrae datos de una tabla dentro de un archivo Markdown. """
    datos_tabla = []
    header_found, separator_found = False, False
    contador_seq_md = 1
    try:
        with open(archivo_md_path, 'r', encoding='utf-8') as f: lines = f.readlines()
    except Exception as e:
        print(f"❌ Error leyendo archivo MD {archivo_md_path.name}: {e}", file=sys.stderr); return []

    for ln, linea in enumerate(lines):
        linea_limpia = linea.strip()
        if not linea_limpia.startswith('|'): continue

        if not header_found:
            lu = linea.upper()
            if ('TITLE' in lu or 'TÍTULO' in lu) and ('TIME' in lu or 'TIEMPO' in lu or 'DURATION' in lu or 'DURACIÓN' in lu):
                header_found = True
            continue

        if header_found and not separator_found:
             if re.match(r'^\s*\|(?:\s*:?[-]+:?\s*\|)+', linea_limpia):
                 separator_found = True
             continue

        if header_found and separator_found:
            match = re.match(REGEX_MARKDOWN_ROW, linea_limpia)
            if match:
                try:
                    titulo = match.group(2).strip()
                    pub_raw = match.group(3).strip()
                    comp_raw = match.group(4).strip()
                    tiempo_str = match.group(5).strip()

                    if not titulo or not tiempo_str: continue

                    tf, ds = parsear_y_formatear_tiempo(tiempo_str)
                    if ds <= 0: continue

                    pub_str = limpiar_participante(pub_raw)
                    comp_str = limpiar_participante(comp_raw)

                    datos_tabla.append({
                        'seq': contador_seq_md, 'title': titulo, 'publisher': pub_str,
                        'composer': comp_str, 'time': tf, 'duration_seconds': ds, 'episode': episodio
                    })
                    contador_seq_md += 1
                except Exception as e_row:
                    print(f"❌ Error procesando fila MD (línea {ln+1}, Ep {episodio}): {e_row}", file=sys.stderr)

    if not header_found or not separator_found:
        print(f"Advertencia: No se encontró tabla de música válida en {archivo_md_path.name}.", file=sys.stderr)
    return datos_tabla

def procesar_cue_sheet(archivo_path_str: str, output_dir: str) -> Dict[str, Any]:
    """ Procesa un único archivo fuente (Excel/MD), extrae datos y genera reporte individual. """
    archivo_path = Path(archivo_path_str)
    nombre_archivo = archivo_path.name
    episodio = DEFAULT_EPISODIO
    datos_tabla = []
    input_is_markdown = False
    output_path = Path(output_dir)
    try: output_path.mkdir(parents=True, exist_ok=True)
    except OSError as e: print(f"❌ Error creando dir salida '{output_dir}': {e}", file=sys.stderr); return {'exito': False, 'mensaje': f"Error dir salida: {e}"}

    try:
        match = re.search(REGEX_EPISODIO, nombre_archivo, re.IGNORECASE)
        if match:
            ep_num_str = next((g for g in match.groups() if g is not None), None)
            if ep_num_str and ep_num_str.isdigit(): episodio = ep_num_str.zfill(3)
            else: print(f"Advertencia: No se encontró número de episodio válido en '{nombre_archivo}'. Usando '{episodio}'.")
        else: print(f"Advertencia: No se detectó número de episodio en '{nombre_archivo}'. Usando por defecto: '{episodio}'.")

        print(f"ℹ️ Procesando '{nombre_archivo}' para Episodio {episodio}")
        file_ext = archivo_path.suffix.lower()

        if file_ext == '.xlsx':
            try:
                wb = openpyxl.load_workbook(archivo_path_str, data_only=True)
                ws = wb.active
                if ws is None: raise ValueError("No se encontró hoja activa.")
                datos_tabla = _procesar_excel_sheet(ws, episodio)
            except Exception as e_excel: raise RuntimeError(f"Error procesando Excel '{nombre_archivo}': {e_excel}") from e_excel
        elif file_ext == '.md':
            input_is_markdown = True
            datos_tabla = _procesar_markdown_sheet(archivo_path, episodio)
        else: raise ValueError(f"Tipo de archivo no soportado: '{nombre_archivo}'.")

        if datos_tabla:
            print(f"✅ Encontradas {len(datos_tabla)} pistas válidas en '{nombre_archivo}'.")
            if not input_is_markdown:
                cont_md_simple = f"# Episodio {episodio}\n\n## Tabla de Pistas (Excel)\n\n"
                cont_md_simple += "| SEQ# | TITLE | PUBLISHER | COMPOSER | TIME (MM:SS) |\n|:----:|:------|:----------|:---------|:------------:|\n"
                for f in datos_tabla:
                    t = (f['title'][:40]+'...') if len(f['title'])>43 else f['title']
                    p = (f['publisher'][:35]+'...') if len(f['publisher'])>38 else f['publisher']
                    c = (f['composer'][:35]+'...') if len(f['composer'])>38 else f['composer']
                    cont_md_simple += f"| {f['seq']} | {t} | {p} | {c} | {f['time']} |\n"
                out_md_simple_path = output_path / MARKDOWN_EPISODIO_FILENAME_FORMAT.format(episodio=episodio)
                try:
                    with open(out_md_simple_path, 'w', encoding='utf-8') as f_mds: f_mds.write(cont_md_simple)
                    print(f"✅ Tabla MD simple generada: {out_md_simple_path}")
                except IOError as e_io_mds: print(f"❌ Error escribiendo MD simple {out_md_simple_path}: {e_io_mds}", file=sys.stderr)

            generar_reporte_estadisticas(datos_tabla, episodio, str(output_path))
            return {'exito': True, 'episodio': episodio, 'datos_tabla': datos_tabla, 'mensaje': f"OK, {len(datos_tabla)} pistas."}
        else:
            print(f"ℹ️ No se encontraron pistas válidas en '{nombre_archivo}'.")
            return {'exito': True, 'episodio': episodio, 'datos_tabla': None, 'mensaje': "OK, sin pistas válidas."}

    except Exception as e_main:
        mensaje = f"Error procesando '{nombre_archivo}' (Ep {episodio}): {e_main}"
        print(f"❌ {mensaje}", file=sys.stderr)
        return {'exito': False, 'episodio': episodio, 'datos_tabla': None, 'mensaje': mensaje}

# =========================
# Lógica Principal y Orquestación
# =========================
def run_processing(files_to_process: List[str], output_dir: str, report_name: str) -> bool:
    """ Orquesta el procesamiento de múltiples archivos y genera reportes globales. """
    if not files_to_process: print("No hay archivos para procesar."); return False

    output_path = Path(output_dir)
    try:
        output_path.mkdir(parents=True, exist_ok=True)
        print(f"Directorio de salida: {output_path.resolve()}")
    except OSError as e_dir: print(f"❌ Error creando dir salida '{output_dir}': {e_dir}", file=sys.stderr); return False

    datos_globales = []
    files_ok_con_datos, files_ok_sin_datos, files_con_errores = 0, 0, 0
    total_files_intentados = len(files_to_process)
    print(f"\nIniciando procesamiento de {total_files_intentados} archivo(s)...")

    for i, file_path_str in enumerate(files_to_process):
        print(f"\n--- Procesando archivo {i+1}/{total_files_intentados}: {os.path.basename(file_path_str)} ---")
        resultado_archivo = procesar_cue_sheet(file_path_str, str(output_path))
        if resultado_archivo:
            if resultado_archivo.get('exito'):
                if resultado_archivo.get('datos_tabla'): files_ok_con_datos += 1; datos_globales.extend(resultado_archivo['datos_tabla'])
                else: files_ok_sin_datos += 1
            else: files_con_errores += 1
        else: files_con_errores += 1; print(f"❌ Fallo inesperado procesando {os.path.basename(file_path_str)}.", file=sys.stderr)

    print("\n" + "="*35 + "\nResumen Procesamiento Individual\n" + "="*35)
    print(f"Total archivos: {total_files_intentados}")
    print(f"  OK con datos: {files_ok_con_datos}")
    if files_ok_sin_datos > 0: print(f"  OK sin datos: {files_ok_sin_datos}")
    if files_con_errores > 0: print(f"  Con errores:  {files_con_errores}")
    print("-"*35)

    if datos_globales:
        print("\n--- Iniciando Generación de Reportes Globales ---")
        # <<< CAMBIO v1.8.0: Pasa report_name a la función >>>
        generar_reportes_globales(datos_globales, output_dir, report_name)
    else:
        print("ℹ️ No se generaron reportes globales (sin datos válidos).")

    print("\n" + "="*35 + "\nProceso de extracción completado.\n" + "="*35)
    return total_files_intentados > 0


# =========================
# Interfaz Gráfica (GUI) - Actualizada v1.8.0 (Título, Escalado)
# =========================
def main_gui():
    """ Inicia la interfaz gráfica de usuario (GUI) para el extractor. """
    if not GUI_ENABLED:
        print("Error: GUI requiere 'ttkbootstrap'. Instala con: pip install ttkbootstrap", file=sys.stderr)
        sys.exit(1)

    warnings = []
    if not MATPLOTLIB_AVAILABLE: warnings.append("Matplotlib no instalado: No se generarán gráficos.")
    if not FPDF2_AVAILABLE: warnings.append("FPDF2 no instalado: No se generará el reporte PDF global.")
    if FPDF2_AVAILABLE and not PIL_AVAILABLE: warnings.append("Pillow no instalado: Gráficos y Logo NO incluidos en PDF.")
    if warnings: messagebox.showwarning("Dependencias Opcionales Faltantes", "\n\n".join(warnings))

    root = ttkb.Window(themename="darkly")

    # <<< CAMBIO v1.8.0: Título de la ventana actualizado >>>
    chart_status = "Gráficos " + ("OK" if MATPLOTLIB_AVAILABLE else "OFF")
    pdf_status = "PDF " + ("OK" if FPDF2_AVAILABLE and PIL_AVAILABLE else ("(sin imgs)" if FPDF2_AVAILABLE else "OFF"))
    root.title(f"Braindog cuesheets v1.8.0 ({chart_status}, {pdf_status})") # Nombre app y versión

    # <<< CAMBIO v1.8.0: Eliminado geometry, duplicado factor de escalado >>>
    # root.geometry("1000x700") # Eliminado para que el escalado controle el tamaño
    try:
        # Duplicar tamaño de UI (afecta ventana, fuentes, widgets)
        root.tk.call('tk', 'scaling', 2.0) # Era 1.5
        print("INFO: Escalado de interfaz gráfica aplicado (x2.0).")
    except Exception as e_scale:
        print(f"Advertencia: No se pudo aplicar el escalado Tkinter: {e_scale}", file=sys.stderr)

    selected_files = []
    selected_label_var = ttkb.StringVar(value="Ningún archivo seleccionado.")
    output_dir_var = ttkb.StringVar(value=str(Path.cwd()))
    # <<< CAMBIO v1.8.0: Usa el nuevo nombre por defecto >>>
    report_name_var = ttkb.StringVar(value=DEFAULT_REPORT_NAME)
    is_processing = ttkb.BooleanVar(value=False)

    def select_files_gui():
        if is_processing.get(): return
        file_types = [("Hojas de Música", "*.xlsx *.md"), ("Todos los archivos", "*.*")]
        filenames = filedialog.askopenfilenames(title="Seleccionar Archivos (.xlsx / .md)", filetypes=file_types)
        if filenames:
            nonlocal selected_files
            valid_files = [str(Path(f).resolve()) for f in filenames if f.lower().endswith(('.xlsx', '.md'))]
            skipped_count = len(filenames) - len(valid_files)
            selected_files = valid_files
            count = len(selected_files)
            if count == 1: selected_label_var.set(f"1 archivo: {Path(selected_files[0]).name}")
            elif count > 1: selected_label_var.set(f"{count} archivos seleccionados.")
            else: selected_label_var.set("Ningún archivo válido (.xlsx/.md) seleccionado.")
            log_text.insert(ttkb.END, f"Seleccionados {count} archivo(s) válido(s).\n")
            if skipped_count > 0: log_text.insert(ttkb.END, f"Omitidos {skipped_count} archivos no soportados.\n", ('warning',))
            log_text.see(ttkb.END)

    def select_output_dir_gui():
        if is_processing.get(): return
        directory = filedialog.askdirectory(title="Seleccionar Directorio de Salida", initialdir=output_dir_var.get() or str(Path.cwd()))
        if directory:
            output_dir_var.set(directory)
            log_text.insert(ttkb.END, f"Directorio de salida: {directory}\n", ('info',))
            log_text.see(ttkb.END)

    def run_processing_gui():
        if is_processing.get(): return
        if not selected_files:
            messagebox.showwarning("Archivos no Seleccionados", "Selecciona archivos (.xlsx o .md).")
            return
        output_dir = output_dir_var.get()
        report_name = report_name_var.get().strip()
        if not output_dir: messagebox.showerror("Directorio Inválido", "Selecciona un directorio de salida."); return
        if not report_name: messagebox.showerror("Nombre Inválido", "El nombre base del reporte no puede estar vacío."); return
        if re.search(r'[<>:"/\\|?*]', report_name):
            messagebox.showerror("Nombre Inválido", "El nombre base contiene caracteres inválidos.\nEvita: <>:\"/\\|?*"); return
        try: Path(output_dir).mkdir(parents=True, exist_ok=True)
        except Exception as e: messagebox.showerror("Directorio Inválido", f"Error con directorio:\n'{output_dir}'\n{e}"); return

        is_processing.set(True)
        select_files_button.config(state=ttkb.DISABLED)
        select_output_button.config(state=ttkb.DISABLED)
        report_name_entry.config(state=ttkb.DISABLED)
        process_button.config(state=ttkb.DISABLED, text="Procesando...")
        root.update_idletasks()

        try:
            charts_path = Path(output_dir) / CHARTS_SUBDIR
            if charts_path.exists() and charts_path.is_dir():
                log_text.insert(ttkb.END, f"Limpiando gráficos anteriores: {charts_path}\n", ('info',))
                shutil.rmtree(charts_path)
        except Exception as e_clean: log_text.insert(ttkb.END, f"Advertencia: No se pudo limpiar dir. gráficos: {e_clean}\n", ('warning',))

        log_text.insert(ttkb.END, "\n" + "="*50 + "\nIniciando proceso...\n", ('info',))
        log_text.insert(ttkb.END, f"Directorio salida: {output_dir}\n", ('info',))
        log_text.insert(ttkb.END, f"Nombre base reporte: '{report_name}'\n", ('info',))
        if not MATPLOTLIB_AVAILABLE: log_text.insert(ttkb.END, "Advertencia: Gráficos desactivados.\n", ('warning',))
        if not FPDF2_AVAILABLE: log_text.insert(ttkb.END, "Advertencia: Reporte PDF desactivado.\n", ('warning',))
        if FPDF2_AVAILABLE and not PIL_AVAILABLE: log_text.insert(ttkb.END, "Advertencia: Gráficos/Logo no irán en PDF (Pillow falta).\n", ('warning',))
        log_text.insert(ttkb.END, "="*50 + "\n\n"); log_text.see(ttkb.END)

        try:
            success = run_processing(selected_files, output_dir, report_name) # Pasa el nombre
            if success:
                log_text.insert(ttkb.END, "\n" + "="*50 + "\nProceso finalizado con éxito.\n" + "="*50 + "\n", ('success',))
                messagebox.showinfo("Proceso Completado", "Reportes generados correctamente.")
            else:
                 if not any(Path(output_dir, f).exists() for f in os.listdir(output_dir) if f.endswith(('.md', '.pdf'))): # Check if any output was generated
                     log_text.insert(ttkb.END, "\n" + "="*50 + "\nProceso finalizado, pero no se generaron reportes (sin datos válidos).\n" + "="*50 + "\n", ('warning',))
                     messagebox.showwarning("Completado sin Datos", "No se encontraron datos válidos en los archivos.")
                 else:
                     log_text.insert(ttkb.END, "\n" + "="*50 + "\nProceso finalizado (estado inesperado).\n" + "="*50 + "\n", ('warning',))
        except Exception as e_gui_run:
            log_text.insert(ttkb.END, f"\n❌ ERROR INESPERADO:\n{e_gui_run}\n", ('error',))
            log_text.insert(ttkb.END, traceback.format_exc() + "\n", ('error',))
            messagebox.showerror("Error Inesperado", f"Error grave:\n{e_gui_run}")
        finally:
            is_processing.set(False)
            select_files_button.config(state=ttkb.NORMAL)
            select_output_button.config(state=ttkb.NORMAL)
            report_name_entry.config(state=ttkb.NORMAL)
            process_button.config(state=ttkb.NORMAL, text=" 4. Procesar Archivos ")
            log_text.see(ttkb.END)

    # --- Construcción de la Interfaz ---
    main_frame = ttkb.Frame(root, padding=15)
    main_frame.pack(fill=ttkb.BOTH, expand=True)

    files_frame = ttkb.Labelframe(main_frame, text=" 1. Archivos de Entrada (.xlsx / .md) ", padding=10)
    files_frame.pack(fill=ttkb.X, pady=(0, 10))
    select_files_button = ttkb.Button(files_frame, text="Seleccionar Archivos", bootstyle="info-outline", command=select_files_gui)
    select_files_button.pack(side=ttkb.LEFT, padx=(0, 10))
    files_label = ttkb.Label(files_frame, textvariable=selected_label_var, anchor=ttkb.W, wraplength=600)
    files_label.pack(side=ttkb.LEFT, fill=ttkb.X, expand=True)

    output_frame = ttkb.Labelframe(main_frame, text=" 2. Directorio de Salida ", padding=10)
    output_frame.pack(fill=ttkb.X, pady=(0, 10))
    select_output_button = ttkb.Button(output_frame, text="Seleccionar Dir.", bootstyle="info-outline", command=select_output_dir_gui)
    select_output_button.pack(side=ttkb.LEFT, padx=(0, 10))
    output_entry = ttkb.Entry(output_frame, textvariable=output_dir_var, state="readonly")
    output_entry.pack(side=ttkb.LEFT, fill=ttkb.X, expand=True)

    report_name_frame = ttkb.Labelframe(main_frame, text=" 3. Nombre Base Reporte Global ", padding=10)
    report_name_frame.pack(fill=ttkb.X, pady=(0, 10))
    report_name_entry = ttkb.Entry(report_name_frame, textvariable=report_name_var)
    report_name_entry.pack(side=ttkb.LEFT, fill=ttkb.X, expand=True, padx=(0, 5))
    report_name_label = ttkb.Label(report_name_frame, text=f"(Generará '{report_name_var.get()}_Global.pdf', etc.)")
    report_name_label.pack(side=ttkb.LEFT)
    def update_example_label(*args):
        name = report_name_var.get().strip() or "Reporte" # Fallback simple
        report_name_label.config(text=f"(Generará '{name}_Global.pdf', etc.)")
    report_name_var.trace_add("write", update_example_label)

    process_frame = ttkb.Frame(main_frame, padding=(0, 10, 0, 10))
    process_frame.pack(fill=ttkb.X)
    process_button = ttkb.Button(process_frame, text=" 4. Procesar Archivos ", bootstyle="success", command=run_processing_gui)
    process_button.pack(pady=5)

    log_frame = ttkb.Labelframe(main_frame, text=" Registro del Proceso ", padding=10)
    log_frame.pack(fill=ttkb.BOTH, expand=True, pady=(10, 0))
    log_text = ScrolledText(log_frame, wrap=ttkb.WORD, height=15, autohide=True)
    log_text.pack(fill=ttkb.BOTH, expand=True, padx=5, pady=5)
    log_text.tag_config('error', foreground=root.style.colors.danger)
    log_text.tag_config('warning', foreground=root.style.colors.warning)
    log_text.tag_config('success', foreground=root.style.colors.success)
    log_text.tag_config('info', foreground=root.style.colors.info)
    log_text.tag_config('stdout', foreground=root.style.colors.fg)

    # Redirección stdout/stderr
    original_stdout, original_stderr = sys.stdout, sys.stderr
    class TextRedirector:
        def __init__(self, widget, stream_type="stdout"):
            self.widget = widget; self.stream_type = stream_type
            self.original_stream = sys.__stdout__ if stream_type == "stdout" else sys.__stderr__
        def write(self, text):
            tag = 'stdout'
            text_lower = text.lower().strip()
            if self.stream_type == 'stderr' or text_lower.startswith(('error', '❌', 'traceback')): tag = 'error'
            elif text_lower.startswith(('advertencia', 'warning', '⚠️')): tag = 'warning'
            elif text_lower.startswith(('✅', 'success', 'ok')): tag = 'success'
            elif text_lower.startswith(('ℹ️', 'info', 'ℹ', 'generando', 'procesando')): tag = 'info'
            try:
                if self.widget.winfo_exists(): self.widget.after(0, self._insert_text, text, (tag,))
            except Exception as e: print(f"Error Redirector ({self.stream_type}): {e}\n{text}", file=self.original_stream)
        def _insert_text(self, text, tags):
            try:
                if self.widget.winfo_exists(): self.widget.insert(ttkb.END, text, tags); self.widget.see(ttkb.END)
            except Exception as e: print(f"Error insertando ({self.stream_type}): {e}\n{text}", file=self.original_stream)
        def flush(self): pass
    sys.stdout = TextRedirector(log_text, "stdout")
    sys.stderr = TextRedirector(log_text, "stderr")

    # Mensaje inicial log
    print("="*50 + f"\nBraindog cuesheets v1.8.0 - GUI\n" + "="*50) # Versión
    print("Bienvenido. Sigue los pasos para generar reportes.")
    print("-"*50)

    try: root.mainloop()
    finally:
        if isinstance(sys.stdout, TextRedirector): sys.stdout = original_stdout
        if isinstance(sys.stderr, TextRedirector): sys.stderr = original_stderr
        print("\nStreams stdout/stderr restaurados.")


# =========================
# Punto de Entrada Principal CLI - Actualizado v1.8.0 (descripción, nombre reporte)
# =========================
def main_cli():
    """ Ejecuta la lógica del extractor desde la línea de comandos (CLI). """
    parser = argparse.ArgumentParser(
        # <<< CAMBIO v1.8.0: Descripción y versión actualizadas >>>
        description="Braindog cuesheets v1.8.0: Procesa CUE Sheets (Excel/MD), genera reportes con gráficos/logo.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="Ejemplo: python extractor_mejorado.py ./cues/ -o ./salida/ --report-name 'Proyecto_X'"
    )
    parser.add_argument('input_paths', nargs='+', help="Ruta(s) a archivos (.xlsx/.md) o directorios.")
    parser.add_argument('--output-dir', '-o', default=".", help="Directorio de salida para reportes (defecto: actual).")
    # <<< CAMBIO v1.8.0: Usa el nuevo nombre por defecto >>>
    parser.add_argument('--report-name', default=DEFAULT_REPORT_NAME, help=f"Nombre base para reportes globales (defecto: {DEFAULT_REPORT_NAME}).")

    args = parser.parse_args()

    # Validación nombre reporte
    report_name = args.report_name.strip()
    if not report_name:
        print(f"❌ Error: --report-name vacío. Usando '{DEFAULT_REPORT_NAME}'.", file=sys.stderr)
        report_name = DEFAULT_REPORT_NAME
    elif re.search(r'[<>:"/\\|?*]', report_name):
        original_name = report_name
        report_name = re.sub(r'[<>:"/\\|?*]', '_', report_name)
        print(f"⚠️ Advertencia: Nombre '{original_name}' contenía inválidos. Se usará '{report_name}'.", file=sys.stderr)

    # Búsqueda archivos entrada
    files_to_process = []
    valid_extensions = {'.xlsx', '.md'}
    processed_paths = set()
    print("Buscando archivos...")
    for input_path_str in args.input_paths:
        input_path = Path(input_path_str)
        try:
            resolved_path = input_path.resolve(strict=False)
            if resolved_path in processed_paths: continue
            if not input_path.exists():
                print(f"⚠️ Ruta no encontrada: '{input_path_str}'. Ignorada.", file=sys.stderr)
                processed_paths.add(resolved_path); continue
            if input_path.is_dir():
                processed_paths.add(resolved_path); print(f"Explorando dir: {resolved_path}")
                found_in_dir = 0
                for item in input_path.iterdir():
                    if item.is_file() and item.suffix.lower() in valid_extensions and not item.name.startswith(('~$', '._')):
                        item_resolved = item.resolve()
                        if item_resolved not in processed_paths:
                            files_to_process.append(str(item_resolved)); processed_paths.add(item_resolved); found_in_dir += 1
                print(f"  -> {found_in_dir} archivo(s) válido(s) encontrado(s).")
            elif input_path.is_file():
                 if input_path.suffix.lower() in valid_extensions and not input_path.name.startswith(('~$', '._')):
                     if resolved_path not in processed_paths:
                         print(f"Añadiendo archivo: {resolved_path}")
                         files_to_process.append(str(resolved_path)); processed_paths.add(resolved_path)
                 else: print(f"⚠️ Omitiendo archivo no soportado: '{input_path_str}'", file=sys.stderr); processed_paths.add(resolved_path)
            else: print(f"⚠️ Ruta ignorada (no es archivo/dir): '{input_path_str}'", file=sys.stderr); processed_paths.add(resolved_path)
        except PermissionError:
             print(f"⚠️ Error permisos en '{input_path_str}'. Ignorada.", file=sys.stderr)
             # Asegura añadir la ruta procesada incluso si falla por permisos
             if 'resolved_path' in locals():
                 processed_paths.add(resolved_path)
        except Exception as e: print(f"Error procesando ruta '{input_path_str}': {e}", file=sys.stderr)

    # Ejecutar procesamiento
    if files_to_process:
        files_to_process.sort(); num_files = len(files_to_process)
        print(f"\nSe procesarán {num_files} archivo(s).")
        print(f"Nombre base reportes globales: '{report_name}'") # Usa nombre validado
        if not MATPLOTLIB_AVAILABLE: print("\n⚠️ ADVERTENCIA: matplotlib no instalado -> Sin gráficos.", file=sys.stderr)
        if not FPDF2_AVAILABLE: print("\n⚠️ ADVERTENCIA: fpdf2 no instalado -> Sin reporte PDF.", file=sys.stderr)
        if FPDF2_AVAILABLE and not PIL_AVAILABLE: print("\n⚠️ ADVERTENCIA: Pillow no instalado -> Gráficos/Logo no irán en PDF.", file=sys.stderr)

        try:
            charts_path = Path(args.output_dir) / CHARTS_SUBDIR
            if charts_path.exists() and charts_path.is_dir():
                print(f"INFO: Limpiando dir. gráficos anterior: {charts_path}")
                shutil.rmtree(charts_path)
        except Exception as e_clean: print(f"Advertencia: No se pudo limpiar dir. gráficos: {e_clean}", file=sys.stderr)

        # <<< CAMBIO v1.8.0: Pasa report_name validado >>>
        run_processing(files_to_process=files_to_process, output_dir=args.output_dir, report_name=report_name)
    else:
        print("\nNo se encontraron archivos válidos (.xlsx o .md) para procesar.")
        sys.exit(1)

# =========================
# Punto de Entrada Principal del Script
# =========================
if __name__ == "__main__":
    if "--gui" in sys.argv:
        if GUI_ENABLED:
            main_gui()
        else:
            print("Error: --gui requiere 'ttkbootstrap'. Instala con: pip install ttkbootstrap", file=sys.stderr)
            sys.exit(1)
    else:
        main_cli()
    import openpyxl
import re
from pathlib import Path
from extractor_mejorado import parsear_y_formatear_tiempo, limpiar_participante

def extract_from_excel(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    datos = []
    for row in ws.iter_rows(min_row=17, values_only=True):
        if not row or not row[3]:
            continue
        _, segundos = parsear_y_formatear_tiempo(row[6])
        composer = limpiar_participante(row[8])
        publisher = limpiar_participante(row[13])
        match = re.search(r'(?:EP|CAP|Episodio)\s*(\d+)', path.stem, re.IGNORECASE)
        episode = match.group(1) if match else '000'
        datos.append({
            'title':    str(row[3]).strip(),
            'duration_seconds': segundos,
            'composer': composer,
            'publisher': publisher,
            'episode':  episode
        })
    return datos

def extract_from_md(path: Path):
    text = path.read_text(encoding='utf-8')
    regex = re.compile(r'^\s*\|\s*(\d+)\s*\|\s*(.*?)\s*\|', re.MULTILINE)
    episodio = path.stem.split('_')[-1]
    datos = []
    for m in regex.finditer(text):
        datos.append({
            'title':           m.group(2),
            'duration_seconds': 0,
            'composer':        'N/A',
            'publisher':       'N/A',
            'episode':         episodio
        })
    return datos
    